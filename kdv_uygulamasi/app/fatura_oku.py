"""GIB portal dokumlerini ve elle hazirlanan fatura listelerini okur.

Uc bicim taninir ve kendiliginden ayirt edilir:

1. **e-Arsiv portal dokumu** — sayfalar "ONAYLI / ONAYSIZ e-arsiv portal" ve
   "e-arsiv portal iptal itiraz". Baslik satirinin ustunde aciklama satirlari
   bulunur. Tarih `20230122170855` bicimindedir.
2. **e-Fatura portal dokumu** — "e-fatura" ve "e-fatura iptal itiraz"
   sayfalari; "Fatura Tipi" sutunu vardir, tarih `2023-01-03` bicimindedir.
3. **Elle hazirlanan fatura dokumu** — her yil bir sayfa; "Malın Cinsi",
   "Yevmiye Tarih/No" sutunlariyla. VKN tasimaz.

Baslik satiri sabit bir satir numarasindan okunmaz; ilk satirlar taranip
taninan sutun adlarini en cok tasiyan satir baslik kabul edilir. Boylece
portal ciktisina bir aciklama satiri eklendiginde okuma bozulmaz.

Iptal/itiraz sayfalari da okunur; orada gecen fatura numaralari isaretlenir,
boylece iptal edilmis bir belge sessizce tarhiyata girmez.
"""
import os
import re
import sys
from datetime import date, datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
for _dizin in (os.path.join(BASE_DIR, "lib"), os.path.join(BASE_DIR, "lib_ek")):
    if _dizin not in sys.path:
        sys.path.append(_dizin)


class FaturaHata(Exception):
    """Kullaniciya gosterilecek okuma hatasi."""


# Normalize edilmis sutun basligi -> ic alan adi.
# Normalizasyon: buyuk harf, Turkce harfler sadelestirilmeden, harf ve rakam
# disindaki her sey atilir. Ayni buyukluk icin farkli yazimlar birlikte durur.
SUTUN_ESLEMESI = {
    "UUID": "uuid",
    "INVOICEUUID": "uuid",
    "FATURANUMARASI": "fatura_no",
    "FATURANO": "fatura_no",
    "FATURATARIHI": "tarih",
    "FATURATARIH": "tarih",
    "GERCEKDUZENLEMETARIHI": "gercek_tarih",
    "FATURATIPI": "fatura_tipi",
    "MUKELLEFVKNTCKN": "duzenleyen_vkn",
    "VKNTCKN": "duzenleyen_vkn",
    "ALICIVKNTCKN": "alici_vkn",
    "ALICIUNVAN": "alici_unvan",
    "ONAYDURUMU": "onay_durumu",
    "VERGIHARICTUTAR": "matrah",
    "TUTAR": "matrah",
    "VERGITUTARI": "kdv",
    "TOPLAMVERGITUTARI": "kdv",
    "KDV": "kdv",
    "TOPLAMTUTAR": "toplam",
    "ODENECEKTUTAR": "odenecek",
    "TOPLAMISKONTO": "iskonto",
    "TOPLAMISKONTOTUTARI": "iskonto",
    "PARABIRIMI": "para_birimi",
    "DOVIZKURU": "kur",
    "MALINCINSI": "mal_cinsi",
    "YEVMIYETARIH": "yevmiye_tarih",
    "YEVMIYETARIHI": "yevmiye_tarih",
    "YEVMIYENO": "yevmiye_no",
    "IPTALITIRAZDURUMU": "iptal_durumu",
    "IPTALITIRAZDURUMU2": "iptal_durumu",
    "TALEPDURUMACIKLAMA": "iptal_aciklama",
    "TALEPDURUMUACIKLAMA": "iptal_aciklama",
}

# Bir satirin baslik satiri sayilmasi icin en az bu kadar taninan sutun gerekir
EN_AZ_TANINAN_SUTUN = 3

# Baslik aranirken bakilacak satir sayisi
BASLIK_ARAMA_DERINLIGI = 12

# Sonu bulmak icin: bu kadar ust uste bos satir gorulurse okuma biter
BOS_SATIR_SINIRI = 25


def _normalize(baslik):
    """Sutun basligini eslestirmeye uygun hale getirir."""
    if baslik is None:
        return ""
    metin = str(baslik).upper()
    # Turkce harfleri ASCII karsiliklarina indir; portal ciktilarinda ayni
    # baslik bazen "TARİHİ" bazen "TARIHI" yazilabiliyor.
    for kaynak, hedef in (("İ", "I"), ("I", "I"), ("Ş", "S"), ("Ğ", "G"),
                          ("Ü", "U"), ("Ö", "O"), ("Ç", "C")):
        metin = metin.replace(kaynak, hedef)
    return re.sub(r"[^A-Z0-9]", "", metin)


def _sayi(deger):
    """Hucredeki tutari sayiya cevirir; bos ya da cozulemezse 0.0."""
    if deger is None or deger == "":
        return 0.0
    if isinstance(deger, (int, float)):
        return float(deger)
    metin = str(deger).strip().replace(" ", "")
    if not metin:
        return 0.0
    # "1.517.800,00" ve "1517800.00" bicimlerinin ikisi de gelebiliyor
    if "," in metin and "." in metin:
        metin = (metin.replace(".", "").replace(",", ".")
                 if metin.rfind(",") > metin.rfind(".")
                 else metin.replace(",", ""))
    elif "," in metin:
        metin = metin.replace(",", ".")
    try:
        return float(metin)
    except ValueError:
        return 0.0


def _tarih(deger):
    """Hucredeki tarihi date nesnesine cevirir; cozulemezse None.

    Uc bicim gelir: hucre zaten tarihse oldugu gibi, "2023-01-03" metni ve
    e-Arsiv portalinin "20230122170855" damgasi.
    """
    if deger is None or deger == "":
        return None
    if isinstance(deger, datetime):
        return deger.date()
    if isinstance(deger, date):
        return deger
    metin = str(deger).strip()
    if not metin:
        return None
    # Sayi olarak okunmus damga: 20230122170855 ya da 20230122
    sadece_rakam = re.sub(r"\D", "", metin)
    if len(sadece_rakam) in (8, 14) and metin.replace(".0", "").isdigit():
        try:
            return date(int(sadece_rakam[0:4]), int(sadece_rakam[4:6]),
                        int(sadece_rakam[6:8]))
        except ValueError:
            return None
    for bicim in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d",
                  "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(metin, bicim).date()
        except ValueError:
            continue
    return None


def _metin(deger):
    if deger is None:
        return ""
    if isinstance(deger, float) and deger.is_integer():
        return str(int(deger))          # VKN'ler sayi olarak okunabiliyor
    return str(deger).strip()


def _baslik_satirini_bul(satirlar):
    """Taninan sutun adini en cok tasiyan satiri baslik kabul eder.

    Doner: (satir_indeksi, {sutun_indeksi: alan_adi}) ya da (None, {}).
    """
    en_iyi, en_iyi_esleme, en_iyi_i = 0, {}, None
    for i, satir in enumerate(satirlar[:BASLIK_ARAMA_DERINLIGI]):
        esleme = {}
        for j, hucre in enumerate(satir):
            alan = SUTUN_ESLEMESI.get(_normalize(hucre))
            if alan and alan not in esleme.values():
                esleme[j] = alan
        if len(esleme) > en_iyi:
            en_iyi, en_iyi_esleme, en_iyi_i = len(esleme), esleme, i
    if en_iyi < EN_AZ_TANINAN_SUTUN:
        return None, {}
    return en_iyi_i, en_iyi_esleme


def _sayfa_satirlari(ws):
    """Sayfayi satir listesine cevirir; sonundaki bos alani atar.

    openpyxl salt-okunur kipte bos sayfalar icin bile 1.048.576 satir
    bildirebiliyor; art arda bos satir gorulunce okuma kesilir.
    """
    satirlar, bos = [], 0
    for satir in ws.iter_rows(values_only=True):
        if satir is None or all(h is None or str(h).strip() == "" for h in satir):
            bos += 1
            if bos >= BOS_SATIR_SINIRI:
                break
            satirlar.append(list(satir) if satir else [])
            continue
        bos = 0
        satirlar.append(list(satir))
    while satirlar and all(h is None or str(h).strip() == "" for h in satirlar[-1]):
        satirlar.pop()
    return satirlar


def _iptal_sayfasi_mi(esleme, sayfa_adi):
    if "iptal_durumu" in esleme.values():
        return True
    return "iptal" in _normalize(sayfa_adi).lower() or "IPTAL" in _normalize(sayfa_adi)


def _kayit_kur(satir, esleme, sayfa_adi, dosya_adi):
    """Bir veri satirini ic bicime cevirir."""
    ham = {}
    for j, alan in esleme.items():
        ham[alan] = satir[j] if j < len(satir) else None

    fatura_no = _metin(ham.get("fatura_no"))
    tarih = _tarih(ham.get("tarih"))
    if not fatura_no and tarih is None:
        return None                       # bos ya da alt bilgi satiri

    matrah = _sayi(ham.get("matrah"))
    kdv = _sayi(ham.get("kdv"))
    toplam = _sayi(ham.get("toplam")) or _sayi(ham.get("odenecek"))
    # Elle hazirlanan dokumde matrah "Tutar" sutunundadir ve KDV ayri gelir;
    # portal dokumunde "Vergi Hariç Tutar" ve "Vergi Tutarı" ayni anlamdadir.
    if not toplam and (matrah or kdv):
        toplam = matrah + kdv
    if not kdv and toplam and matrah:
        kdv = toplam - matrah

    yevmiye_tarih = _tarih(ham.get("yevmiye_tarih"))
    return {
        "fatura_no": fatura_no,
        "uuid": _metin(ham.get("uuid")),
        "tarih": tarih.isoformat() if tarih else "",
        "yevmiye_tarih": yevmiye_tarih.isoformat() if yevmiye_tarih else "",
        "yevmiye_no": _metin(ham.get("yevmiye_no")),
        "duzenleyen_vkn": _metin(ham.get("duzenleyen_vkn")),
        "alici_vkn": _metin(ham.get("alici_vkn")),
        "alici_unvan": _metin(ham.get("alici_unvan")),
        "mal_cinsi": _metin(ham.get("mal_cinsi")),
        "fatura_tipi": _metin(ham.get("fatura_tipi")),
        "onay_durumu": _metin(ham.get("onay_durumu")),
        "matrah": round(matrah, 2),
        "kdv": round(kdv, 2),
        "toplam": round(toplam, 2),
        "kaynak": "%s / %s" % (dosya_adi, sayfa_adi),
    }


def _iptal_anahtarlari(satirlar, esleme):
    """Iptal/itiraz sayfasindaki fatura numaralarini ve uuid'leri toplar."""
    anahtarlar = set()
    for satir in satirlar:
        ham = {alan: (satir[j] if j < len(satir) else None)
               for j, alan in esleme.items()}
        for alan in ("fatura_no", "uuid"):
            deger = _metin(ham.get(alan))
            if deger:
                anahtarlar.add(deger.upper())
    return anahtarlar


def dosya_oku(yol, dosya_adi=None):
    """Bir xlsx dosyasindaki butun sayfalari okur.

    Doner: {"faturalar": [...], "iptaller": n, "sayfalar": [...],
            "uyarilar": [...]}
    """
    import openpyxl

    dosya_adi = dosya_adi or os.path.basename(yol)
    try:
        wb = openpyxl.load_workbook(yol, read_only=True, data_only=True)
    except Exception as exc:                                    # noqa: BLE001
        raise FaturaHata("Excel dosyası açılamadı: %s" % exc)

    faturalar, sayfa_bilgisi, uyarilar = [], [], []
    iptal_anahtarlari = set()
    try:
        for sayfa_adi in wb.sheetnames:
            satirlar = _sayfa_satirlari(wb[sayfa_adi])
            baslik_i, esleme = _baslik_satirini_bul(satirlar)
            if baslik_i is None:
                sayfa_bilgisi.append({"ad": sayfa_adi, "adet": 0,
                                      "durum": "başlık satırı tanınmadı"})
                continue
            veri = satirlar[baslik_i + 1:]
            if _iptal_sayfasi_mi(esleme, sayfa_adi):
                yeni = _iptal_anahtarlari(veri, esleme)
                iptal_anahtarlari |= yeni
                sayfa_bilgisi.append({"ad": sayfa_adi, "adet": len(yeni),
                                      "durum": "iptal/itiraz listesi"})
                continue
            okunan = 0
            for satir in veri:
                kayit = _kayit_kur(satir, esleme, sayfa_adi, dosya_adi)
                if kayit:
                    faturalar.append(kayit)
                    okunan += 1
            sayfa_bilgisi.append({"ad": sayfa_adi, "adet": okunan,
                                  "durum": "okundu"})
    finally:
        wb.close()

    # Iptal edilmis belgeleri isaretle (silme: kullanici gorsun ve karar versin)
    iptal_sayisi = 0
    for f in faturalar:
        anahtar = (f["fatura_no"] or "").upper()
        uuid = (f["uuid"] or "").upper()
        f["iptal"] = bool((anahtar and anahtar in iptal_anahtarlari)
                          or (uuid and uuid in iptal_anahtarlari))
        if f["iptal"]:
            iptal_sayisi += 1

    if not faturalar:
        uyarilar.append("Dosyada tanınan biçimde fatura satırı bulunamadı.")
    if iptal_sayisi:
        uyarilar.append("%d fatura iptal/itiraz listesinde bulundu ve işaretlendi."
                        % iptal_sayisi)
    return {"faturalar": faturalar, "iptaller": iptal_sayisi,
            "sayfalar": sayfa_bilgisi, "uyarilar": uyarilar}
