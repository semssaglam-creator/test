"""KDV inceleme calismasini Excel'e aktarir.

Uretilen dosya, elde kullanilan calismanin duzenini korur:
    - Her yil icin bir sayfa: elestirili beyan / elestiri girisi / ham beyan
    - Uc ozet tablo: elestirili, beyan edilen ve ikisinin farki
    - Cok yilli dosyalarda ek bir "Ozet" sayfasi: tum donemler tek tabloda

Formuller
---------
Dosya "canli" uretilir: yalnizca girdiler (beyan edilen degerler ve inceleme
tespitleri) sabit sayidir. Elestirili beyan, ozet tablolar, fark tablosu ve
tarhiyat ozeti Excel formulleriyle yazilir. Boylece Excel'de bir beyan rakami
ya da bir tespit degistirildiginde devir zinciri butun donemlere -- yil
sinirini asarak, sayfalar arasi baglantiyla -- kendiliginden yansir.

Formullerin karsiligi hesap.py'deki `_donem_hesapla` fonksiyonudur.
"""
import os
import sys

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LIB_DIR = os.path.join(BASE_DIR, "lib")
if LIB_DIR not in sys.path:
    sys.path.insert(0, LIB_DIR)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .satirlar import (AYLAR_BUYUK, BEYAN_SATIRLARI, BEYAN_TOPLAM_TURLERI,
                       ELESTIRI_ALANLARI, OZET_KOLONLARI, TARHIYAT_KOLONLARI,
                       TOPLAM_TURLERI, gorunen_tarhiyat_kolonlari,
                       varsayilan_kdv_orani)

FONT = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_BASLIK = Font(name="Calibri", size=11, bold=True)
FONT_FARK = Font(name="Calibri", size=10, bold=True, color="C00000")
FONT_NOT = Font(name="Calibri", size=9, italic=True, color="5A6B7D")

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SOL = Alignment(horizontal="left", vertical="center", wrap_text=True)
ORTA = Alignment(horizontal="center", vertical="center", wrap_text=True)
SAG = Alignment(horizontal="right", vertical="center")

DOLGU_BASLIK = PatternFill("solid", fgColor="D9E1F2")
DOLGU_ELESTIRI = PatternFill("solid", fgColor="FFF2CC")
DOLGU_FARK = PatternFill("solid", fgColor="FCE4D6")
DOLGU_TOPLAM = PatternFill("solid", fgColor="E2EFDA")
DOLGU_HESAP = PatternFill("solid", fgColor="F2F5F9")
# Tarhiyat tablosunun vurgu renkleri (elde kullanilan tablodaki gibi)
DOLGU_SARI = PatternFill("solid", fgColor="FFFF99")
DOLGU_YESIL = PatternFill("solid", fgColor="C6E0B4")

SAYI_BICIMI = "#,##0.00"

# Elestirili blogun satir duzeni. Her yil sayfasinda ayni sirada ve ayni satir
# numaralarinda durur; sayfalar arasi devir baglantisi buna dayanir.
ELESTIRILI_ALANLARI = [
    ("matrah", "KDV Matrahı Toplamı"),
    ("hesaplanan", "Hesaplanan KDV"),
    ("ilave_edilecek", "İlave Edilecek KDV"),
    ("toplam_kdv", "Toplam KDV"),
    ("onceki_devir", "Önceki Dönemden Devreden İnd. KDV"),
    ("bu_donem_indirim", "Bu Döneme Ait İndirilecek KDV"),
    ("diger_indirim", "Bu Döneme Ait İnd. Diğer KDV"),
    ("indirimler", "İndirimler Toplamı"),
    ("odenecek", "Ödenmesi Gereken KDV"),
    ("sonraki_devir", "Sonraki Döneme Devreden KDV"),
    ("iade", "İade Edilmesi Gereken KDV"),
    ("tecil_edilecek", "Tecil Edilecek KDV"),
]

# Ozet kolonlarinin beyan blogundaki karsiliklari (tek satir ya da iki satirin
# toplami). Ozet tablolarin beyan blogu bu eslemeyle formule cevrilir.
OZET_BEYAN_KAYNAGI = {
    "matrah": ["matrah_toplami"],
    "toplam_kdv": ["toplam_kdv"],
    "onceki_devir": ["onceki_donem_devreden"],
    "bu_donem_indirim_toplam": ["bu_donem_indirilecek", "diger_indirimler_toplami"],
    "indirimler": ["indirimler_toplami"],
    "odenecek": ["odenmesi_gereken_kdv"],
    "sonraki_devir": ["sonraki_donem_devreden"],
    "iade": ["iade_edilmesi_gereken_kdv"],
}

# Ozet kolonlarinin elestirili blogundaki karsiliklari
OZET_ELESTIRILI_KAYNAGI = {
    "matrah": ["matrah"],
    "toplam_kdv": ["toplam_kdv"],
    "onceki_devir": ["onceki_devir"],
    "bu_donem_indirim_toplam": ["bu_donem_indirim", "diger_indirim"],
    "indirimler": ["indirimler"],
    "odenecek": ["odenecek"],
    "sonraki_devir": ["sonraki_devir"],
    "iade": ["iade"],
}

NOT_TOPLAM = ("TOPLAM sütunu / Toplam satırı sütunun düz toplamıdır. Devir "
              "satırları ile indirimler toplamı stok kalemi taşır; düz toplamları "
              "muhasebe anlamında bir yıl devri değildir, 12 aylık karşılaştırma "
              "içindir. Açılış devri için Ocak, kapanış devri için son ay sütununa "
              "bakınız.")


def _yaz(ws, satir, sutun, deger, font=FONT, hiza=SAG, dolgu=None, bicim=SAYI_BICIMI):
    h = ws.cell(satir, sutun, deger)
    h.font = font
    h.alignment = hiza
    h.border = BORDER
    if dolgu:
        h.fill = dolgu
    formul = isinstance(deger, str) and deger.startswith("=")
    if bicim and (isinstance(deger, (int, float)) or formul):
        h.number_format = bicim
    return h


def _not_satiri(ws, satir, metin, genislik):
    _yaz(ws, satir, 1, metin, FONT_NOT, SOL, bicim=None)
    ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=genislik)
    ws.row_dimensions[satir].height = 26
    return satir + 1


def _baslik_satiri(ws, satir, metin, genislik=14):
    h = ws.cell(satir, 1, metin)
    h.font = FONT_BASLIK
    h.alignment = SOL
    ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=genislik)
    return satir + 1


def _ozet_tablo(ws, satir, baslik, donemler, blok, dolgu=None, fark_mi=False,
                formul=None):
    """Ay x kolon ozet tablosu yazar (elestirili / beyan / fark bloklari).

    formul: (donem_sirasi, alan) -> formul metni ya da None. None donerse
    hesaplanmis deger sabit olarak yazilir.

    Doner: (sonraki_satir, ilk_veri_satiri)
    """
    satir = _baslik_satiri(ws, satir, baslik, len(OZET_KOLONLARI) + 2)
    _yaz(ws, satir, 1, "Dönem", FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
    for i, (_alan, etiket) in enumerate(OZET_KOLONLARI):
        _yaz(ws, satir, i + 2, etiket, FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
    satir += 1

    ilk_veri = satir
    for sira, d in enumerate(donemler):
        _yaz(ws, satir, 1, f"{d['yil']}/{d['ay_adi']}", FONT, SOL, dolgu, bicim=None)
        for i, (alan, _etiket) in enumerate(OZET_KOLONLARI):
            deger = d[blok].get(alan, 0.0)
            font = FONT_FARK if (fark_mi and abs(deger) > 0.005) else FONT
            icerik = formul(sira, alan) if formul else None
            _yaz(ws, satir, i + 2, icerik if icerik else deger, font, SAG, dolgu)
        satir += 1

    # --- Toplam satiri: her sutunun duz toplami
    son_veri = satir - 1
    _yaz(ws, satir, 1, "Toplam", FONT_BOLD, SOL, DOLGU_TOPLAM, bicim=None)
    for i, (alan, _e) in enumerate(OZET_KOLONLARI):
        harf = get_column_letter(i + 2)
        _yaz(ws, satir, i + 2, f"=SUM({harf}{ilk_veri}:{harf}{son_veri})",
             FONT_BOLD, SAG, DOLGU_TOPLAM)
    satir += 1

    # --- Stok kalemleri icin referans satiri
    _yaz(ws, satir, 1, "Açılış / kapanış", FONT, SOL, bicim=None)
    for i, (alan, _e) in enumerate(OZET_KOLONLARI):
        harf = get_column_letter(i + 2)
        tur = TOPLAM_TURLERI.get(alan)
        if tur == "acilis":
            icerik = f"={harf}{ilk_veri}"
        elif tur == "kapanis":
            icerik = f"={harf}{son_veri}"
        elif tur == "duzeltilmis":
            # Devir bir kez sayilmis hali: acilis devri + donem indirimleri
            onc = get_column_letter([a for a, _x in OZET_KOLONLARI].index("onceki_devir") + 2)
            bu = get_column_letter(
                [a for a, _x in OZET_KOLONLARI].index("bu_donem_indirim_toplam") + 2)
            icerik = f"={onc}{ilk_veri}+SUM({bu}{ilk_veri}:{bu}{son_veri})"
        else:
            icerik = ""
        _yaz(ws, satir, i + 2, icerik, FONT, SAG, bicim=SAYI_BICIMI if icerik else None)
    satir += 1

    satir = _not_satiri(
        ws, satir,
        "Toplam satırı her sütunun düz toplamıdır. Devir sütunları ile "
        "indirimler toplamı stok kalemi taşır: devir bir dönemden diğerine "
        "aktarıldığı için düz toplamda tekrar tekrar sayılır. Alttaki satırda "
        "serinin açılış devri, kapanış devri ve indirimlerin devir bir kez "
        "sayılmış hali verilmiştir.", len(OZET_KOLONLARI) + 1)
    return satir + 1, ilk_veri


def _yil_sayfasi(wb, yil_kaydi, donemler, onceki_yil):
    """Bir yilin sayfasini yazar.

    onceki_yil: seride bu yildan once gelen yilin (yil, ay_sayisi) ikilisi ya
    da None. Ocak ayinin onceki devri buradan -- sayfalar arasi baglantiyla --
    alinir; boylece devir zinciri Excel'de de yil sinirini asar.

    Doner: sayfa duzeni sozlugu (satir numaralari, ay sayisi).
    """
    yil = yil_kaydi["yil"]
    ws = wb.create_sheet(str(yil))
    ws.column_dimensions["A"].width = 48
    for i in range(2, 15):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "B2"

    yil_donemleri = [d for d in donemler if d["yil"] == yil]
    ay_sayisi = max(len(yil_donemleri), 1)
    sut = [get_column_letter(i + 2) for i in range(ay_sayisi)]
    son_harf = sut[-1]

    satir = 1
    _yaz(ws, satir, 1, f"{yil} Dönemi", FONT_BASLIK, SOL, DOLGU_BASLIK, bicim=None)
    for i, ay in enumerate(AYLAR_BUYUK):
        _yaz(ws, satir, i + 2, ay, FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
    _yaz(ws, satir, 14, "TOPLAM", FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
    satir += 1

    # --- Elestirili blok: once yerler ayrilir, formuller beyan ve elestiri
    # satir numaralari belli olduktan sonra doldurulur
    satir = _baslik_satiri(ws, satir, "OLMASI GEREKEN (YENİDEN HESAPLANMIŞ) BEYAN")
    x = {}
    for alan, etiket in ELESTIRILI_ALANLARI:
        _yaz(ws, satir, 1, etiket, FONT, SOL, bicim=None)
        x[alan] = satir
        satir += 1
    x_not = satir
    satir += 2

    # --- Elestiri (tespit) girisi: kullanicinin girdigi degerler, sabit
    satir = _baslik_satiri(ws, satir, "İNCELEME TESPİTLERİ (ELEŞTİRİ)")
    elestiri = yil_kaydi.get("elestiri") or {}
    e = {}
    for alan, etiket in ELESTIRI_ALANLARI:
        _yaz(ws, satir, 1, etiket, FONT, SOL, DOLGU_ELESTIRI, bicim=None)
        dizi = elestiri.get(alan) or [0.0] * 12
        for i in range(ay_sayisi):
            _yaz(ws, satir, i + 2, float(dizi[i] or 0), FONT, SAG, DOLGU_ELESTIRI)
        _yaz(ws, satir, 14, f"=SUM(B{satir}:{son_harf}{satir})",
             FONT_BOLD, SAG, DOLGU_TOPLAM)
        e[alan] = satir
        satir += 1
    _yaz(ws, satir, 1, "Uygulanan KDV Oranı (%)", FONT, SOL, DOLGU_ELESTIRI, bicim=None)
    oranlar = elestiri.get("kdv_orani") or [None] * 12
    oran_satir = satir
    for i in range(ay_sayisi):
        oran = oranlar[i] if i < len(oranlar) and oranlar[i] is not None \
            else varsayilan_kdv_orani(yil, i + 1)
        _yaz(ws, satir, i + 2, float(oran), FONT, SAG, DOLGU_ELESTIRI, bicim="0.##")
    satir += 1
    satir = _not_satiri(
        ws, satir,
        "Sarı alanlar ve beyan blokundaki değerler girdidir; diğer bütün "
        "tutarlar formülle hesaplanır. Buradaki bir tespiti değiştirdiğinizde "
        "devir zinciri izleyen bütün dönemlere (yıl sınırını aşarak) yansır.", 14)
    satir += 1

    # --- Ham beyan verisi: sistemden alinan degerler, sabit
    satir = _baslik_satiri(ws, satir, "BEYAN EDİLEN DEĞERLER (SİSTEMDEN ALINAN)")
    _yaz(ws, satir, 1, "DÖNEMİ", FONT_BOLD, SOL, DOLGU_BASLIK, bicim=None)
    for i, ay in enumerate(AYLAR_BUYUK):
        _yaz(ws, satir, i + 2, ay, FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
    _yaz(ws, satir, 14, "TOPLAM", FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
    satir += 1
    beyan = yil_kaydi.get("beyan") or {}
    b = {}
    for kod, etiket, baslik in BEYAN_SATIRLARI:
        if baslik:
            _yaz(ws, satir, 1, etiket, FONT_BOLD, SOL, DOLGU_BASLIK, bicim=None)
            for i in range(1, 14):
                _yaz(ws, satir, i + 1, "", FONT, SAG, DOLGU_BASLIK, bicim=None)
            satir += 1
            continue
        _yaz(ws, satir, 1, etiket, FONT, SOL, bicim=None)
        dizi = beyan.get(kod) or [0.0] * 12
        for i in range(12):
            _yaz(ws, satir, i + 2, float(dizi[i] or 0))
        hucre = _yaz(ws, satir, 14, f"=SUM(B{satir}:M{satir})",
                     FONT_BOLD, SAG, DOLGU_TOPLAM)
        if BEYAN_TOPLAM_TURLERI.get(kod):
            hucre.font = Font(name="Calibri", size=10, bold=True, color="8A6D1B")
        b[kod] = satir
        satir += 1
    satir = _not_satiri(ws, satir, NOT_TOPLAM, 14)
    satir += 1

    # --- Elestirili blok formulleri (hesap.py: _donem_hesapla) ---
    otomatik = elestiri.get("hesaplanan_otomatik") or [True] * 12
    if onceki_yil:
        onc_yil, onc_ay = onceki_yil
        devir_kaynagi = f"'{onc_yil}'!{get_column_letter(onc_ay + 1)}{x['sonraki_devir']}"
    else:
        devir_kaynagi = None
    pin = yil_kaydi.get("devir_baslangic")
    pin = None if pin in (None, "") else float(pin)

    for i in range(ay_sayisi):
        c = sut[i]
        fark = f"({c}{x['toplam_kdv']}-{c}{x['indirimler']})"
        yuklenilen = (f"MAX({c}{b['tam_istisna_yuklenilen']}+{c}{b['diger_iade_kdv']}"
                      f"-{c}{e['yuklenilen_cikar']},0)")
        if i == 0:
            if pin is not None:
                onceki = f"{pin!r}"
            elif devir_kaynagi:
                onceki = devir_kaynagi
            else:
                onceki = f"{c}{b['onceki_donem_devreden']}"
        else:
            onceki = f"{sut[i - 1]}{x['sonraki_devir']}"

        if i < len(otomatik) and otomatik[i] is False:
            hesaplanan_ilave = f"{c}{e['hesaplanan_kdv_ilave']}"
        else:
            hesaplanan_ilave = f"ROUND({c}{e['matrah_ilave']}*{c}{oran_satir}/100,2)"

        formuller = {
            "matrah": f"{c}{b['matrah_toplami']}+{c}{e['matrah_ilave']}",
            "hesaplanan": f"{c}{b['hesaplanan_kdv']}+{hesaplanan_ilave}",
            "ilave_edilecek": f"{c}{b['ilave_edilecek_kdv']}",
            "toplam_kdv": f"{c}{x['hesaplanan']}+{c}{x['ilave_edilecek']}",
            "onceki_devir": f"{onceki}-{c}{e['devir_cikar']}",
            "bu_donem_indirim": f"{c}{b['bu_donem_indirilecek']}-{c}{e['indirim_cikar']}",
            "diger_indirim": f"{c}{b['diger_indirimler_toplami']}",
            "indirimler": (f"{c}{x['onceki_devir']}+{c}{x['bu_donem_indirim']}"
                           f"+{c}{x['diger_indirim']}"),
            "tecil_edilecek": (f"IF({fark}>0,MIN({fark},{c}{b['tecil_edilebilir_kdv']}),0)"),
            "odenecek": f"IF({fark}>0,{fark}-{c}{x['tecil_edilecek']},0)",
            "iade": f"IF({fark}>0,0,MIN(-{fark},{yuklenilen}))",
            "sonraki_devir": f"IF({fark}>0,0,-{fark}-{c}{x['iade']})",
        }
        for alan, _etiket in ELESTIRILI_ALANLARI:
            _yaz(ws, x[alan], i + 2, f"=ROUND({formuller[alan]},2)",
                 FONT, SAG, DOLGU_HESAP)

    for alan, _etiket in ELESTIRILI_ALANLARI:
        hucre = _yaz(ws, x[alan], 14, f"=SUM(B{x[alan]}:{son_harf}{x[alan]})",
                     FONT_BOLD, SAG, DOLGU_TOPLAM)
        if TOPLAM_TURLERI.get(alan):
            hucre.font = Font(name="Calibri", size=10, bold=True, color="8A6D1B")
    _not_satiri(ws, x_not, NOT_TOPLAM, 14)

    # --- Uc ozet tablo (yil ici) ---
    def el_formulu(sira, alan):
        c = sut[sira]
        return "=" + "+".join(f"{c}{x[k]}" for k in OZET_ELESTIRILI_KAYNAGI[alan])

    def by_formulu(sira, alan):
        c = sut[sira]
        return "=" + "+".join(f"{c}{b[k]}" for k in OZET_BEYAN_KAYNAGI[alan])

    satir, el_ilk = _ozet_tablo(ws, satir, "ÖZET — OLMASI GEREKEN", yil_donemleri,
                                "elestirili", formul=el_formulu)
    satir, by_ilk = _ozet_tablo(ws, satir, "ÖZET — BEYAN EDİLEN", yil_donemleri,
                                "beyan", formul=by_formulu)

    def fark_formulu(sira, alan):
        j = [a for a, _e in OZET_KOLONLARI].index(alan) + 2
        h = get_column_letter(j)
        return f"={h}{el_ilk + sira}-{h}{by_ilk + sira}"

    _ozet_tablo(ws, satir, "FARK (OLMASI GEREKEN − BEYAN)", yil_donemleri, "fark",
                DOLGU_FARK, fark_mi=True, formul=fark_formulu)
    return {"yil": yil, "ay_sayisi": ay_sayisi, "x": x, "b": b, "e": e}


def _ozet_sayfasi(wb, inceleme, sonuc, bulgular, duzen):
    ws = wb.create_sheet("Özet", 0)
    ws.column_dimensions["A"].width = 20
    for i in range(2, 11):
        ws.column_dimensions[get_column_letter(i)].width = 16

    satir = 1
    _yaz(ws, satir, 1, "KDV İNCELEME ÇALIŞMASI", FONT_BASLIK, SOL, DOLGU_BASLIK, bicim=None)
    ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=9)
    satir += 2
    for etiket, deger in (("Mükellef", inceleme.get("ad_unvan", "")),
                          ("VKN/TCKN", inceleme.get("vkn_tckn", "")),
                          ("Dosya", inceleme.get("ad", ""))):
        _yaz(ws, satir, 1, etiket, FONT_BOLD, SOL, bicim=None)
        _yaz(ws, satir, 2, deger, FONT, SOL, bicim=None)
        satir += 1
    satir += 1

    donemler = sonuc["donemler"]

    def el_formulu(sira, alan):
        d = donemler[sira]
        sayfa, c = f"'{d['yil']}'!", get_column_letter(d["ay"] + 1)
        return "=" + "+".join(f"{sayfa}{c}{duzen['x'][k]}"
                              for k in OZET_ELESTIRILI_KAYNAGI[alan])

    def by_formulu(sira, alan):
        d = donemler[sira]
        sayfa, c = f"'{d['yil']}'!", get_column_letter(d["ay"] + 1)
        return "=" + "+".join(f"{sayfa}{c}{duzen['b'][k]}"
                              for k in OZET_BEYAN_KAYNAGI[alan])

    satir, el_ilk = _ozet_tablo(ws, satir, "TÜM DÖNEMLER — OLMASI GEREKEN", donemler,
                                "elestirili", formul=el_formulu)
    satir, by_ilk = _ozet_tablo(ws, satir, "TÜM DÖNEMLER — BEYAN EDİLEN", donemler,
                                "beyan", formul=by_formulu)

    def fark_formulu(sira, alan):
        h = get_column_letter([a for a, _e in OZET_KOLONLARI].index(alan) + 2)
        return f"={h}{el_ilk + sira}-{h}{by_ilk + sira}"

    satir, _ = _ozet_tablo(ws, satir, "TÜM DÖNEMLER — FARK", donemler, "fark",
                           DOLGU_FARK, fark_mi=True, formul=fark_formulu)

    if bulgular:
        satir = _baslik_satiri(ws, satir, "BEYAN TUTARLILIK BULGULARI", 9)
        for bulgu in bulgular:
            _yaz(ws, satir, 1, bulgu["donem"], FONT_BOLD, SOL, DOLGU_FARK, bicim=None)
            h = _yaz(ws, satir, 2, bulgu["mesaj"], FONT, SOL, DOLGU_FARK, bicim=None)
            h.alignment = SOL
            ws.merge_cells(start_row=satir, start_column=2, end_row=satir, end_column=9)
            satir += 1
    return ws


def _tarhiyat_sayfasi(wb, inceleme, sonuc, duzen, ziya_kati=1):
    """Elde kullanilan tarhiyat ozeti tablosunun karsiligi (formullu).

    Tutar tasimayan iade sutunlari yazilmaz: bos sutunlar tabloyu Word'e
    yapistirildiginda sayfaya sigmayacak kadar genisletiyordu. Formuller de
    bu yuzden sabit sutun harfleriyle degil, gosterilen sutunlarin kendi
    harfleriyle kurulur.
    """
    kolonlar = gorunen_tarhiyat_kolonlari(sonuc["donemler"])
    ws = wb.create_sheet("Tarhiyat Özeti")
    ws.column_dimensions["A"].width = 14
    for i in range(2, len(kolonlar) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 16

    satir = 1
    _yaz(ws, satir, 1, "TARHİYAT ÖZETİ", FONT_BASLIK, SOL, DOLGU_BASLIK, bicim=None)
    ws.merge_cells(start_row=satir, start_column=1, end_row=satir,
                   end_column=len(kolonlar) + 1)
    satir += 2

    # Gruplanmis ust baslik
    _yaz(ws, satir, 1, "Dönemi", FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
    ws.merge_cells(start_row=satir, start_column=1, end_row=satir + 1, end_column=1)
    i = 0
    while i < len(kolonlar):
        j = i
        while (j + 1 < len(kolonlar)
               and kolonlar[j + 1][0] == kolonlar[i][0]):
            j += 1
        dolgu = _tarhiyat_dolgu(kolonlar[i][3]) or DOLGU_BASLIK
        # Once tum hucreler bicimlendirilir, birlestirme en sonda yapilir:
        # birlestirilen hucrelere sonradan yazilamaz
        for k in range(i + 2, j + 3):
            _yaz(ws, satir, k, "", FONT_BOLD, ORTA, dolgu, bicim=None)
        _yaz(ws, satir, i + 2, kolonlar[i][0], FONT_BOLD, ORTA, dolgu, bicim=None)
        if j > i:
            ws.merge_cells(start_row=satir, start_column=i + 2, end_row=satir, end_column=j + 2)
        i = j + 1
    satir += 1
    for idx, (grup, _kod, etiket, vurgu) in enumerate(kolonlar):
        dolgu = _tarhiyat_dolgu(vurgu) or DOLGU_BASLIK
        _yaz(ws, satir, idx + 2, "" if etiket == grup else etiket, FONT_BOLD, ORTA,
             dolgu, bicim=None)
    satir += 1

    x, b = duzen["x"], duzen["b"]
    ilk_veri = satir
    onceki_yil = None
    for d in sonuc["donemler"]:
        if d["yil"] != onceki_yil:
            _yaz(ws, satir, 1, f"{d['yil']} Dönemi", FONT_BOLD, SOL, DOLGU_BASLIK, bicim=None)
            for k in range(2, len(kolonlar) + 2):
                _yaz(ws, satir, k, "", FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
            satir += 1
            onceki_yil = d["yil"]
        s, c = f"'{d['yil']}'!", get_column_letter(d["ay"] + 1)
        # Gosterilen sutunlarin harfleri; gizlenen bir sutuna atif gerekirse
        # yerine 0 konur (gizlenmis olmasi butun donemlerde sifir demektir).
        harf = {kod: get_column_letter(idx + 2)
                for idx, (_g, kod, _e, _v) in enumerate(kolonlar)}
        H = lambda kod: ("%s%d" % (harf[kod], satir)) if kod in harf else "0"
        formuller = {
            "odenecek_olmasi_gereken": f"={s}{c}{x['odenecek']}",
            "odenecek_beyan": f"={s}{c}{b['odenmesi_gereken_kdv']}",
            "resen_tarhi_gereken":
                f"=MAX({H('odenecek_olmasi_gereken')}-{H('odenecek_beyan')},0)",
            "vergi_ziyai_cezasi": f"={H('resen_tarhi_gereken')}*{int(ziya_kati)}",
            "iade_olmasi_gereken": f"={s}{c}{x['iade']}",
            "iade_beyan": f"={s}{c}{b['iade_edilmesi_gereken_kdv']}",
            "aranmasi_gereken":
                f"=MAX({H('iade_beyan')}-{H('iade_olmasi_gereken')},0)",
            "resen_toplam": f"={H('resen_tarhi_gereken')}+{H('aranmasi_gereken')}",
            # Elestirili ihracat iadesi: tecil edilebilir - tecil edilecek
            "ihracat_iade_olmasi_gereken": (f"={s}{c}{b['tecil_edilebilir_kdv']}"
                                            f"-{s}{c}{x['tecil_edilecek']}"),
            "ihracat_iade_beyan": f"={s}{c}{b['ihracat_tecil_edilemeyen']}",
            "haksiz_iade":
                f"=MAX({H('ihracat_iade_beyan')}-{H('ihracat_iade_olmasi_gereken')},0)",
            "toplam_fark": (f"={H('resen_tarhi_gereken')}+{H('aranmasi_gereken')}"
                            f"+{H('haksiz_iade')}"),
        }
        _yaz(ws, satir, 1, d["ay_adi"], FONT, SOL, bicim=None)
        for idx, (_g, kod, _e, vurgu) in enumerate(kolonlar):
            font = FONT_FARK if vurgu == "vurgu1" else (FONT_BOLD if vurgu == "vurgu2" else FONT)
            _yaz(ws, satir, idx + 2, formuller[kod], font, SAG, _tarhiyat_dolgu(vurgu))
        satir += 1

    _yaz(ws, satir, 1, "Toplam", FONT_BOLD, SOL, DOLGU_TOPLAM, bicim=None)
    for idx, (_g, _kod, _e, vurgu) in enumerate(kolonlar):
        sutun = get_column_letter(idx + 2)
        _yaz(ws, satir, idx + 2, f"=SUM({sutun}{ilk_veri}:{sutun}{satir - 1})",
             FONT_BOLD, SAG, _tarhiyat_dolgu(vurgu) or DOLGU_TOPLAM)
    satir += 2

    t = sonuc.get("tarhiyat_toplami") or {}
    for etiket, kod in (("Mükellefin fazladan beyan ettiği ödenecek KDV", "fazla_beyan_odenecek"),
                        ("Mükellefin talep etmediği iade tutarı", "eksik_talep_iade")):
        if t.get(kod, 0):
            _yaz(ws, satir, 1, etiket, FONT, SOL, DOLGU_FARK, bicim=None)
            ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=4)
            _yaz(ws, satir, 5, t[kod], FONT_BOLD, SAG, DOLGU_FARK)
            _yaz(ws, satir, 6, "(tarhiyata dahil edilmemiştir)", FONT, SOL, DOLGU_FARK,
                 bicim=None)
            satir += 1
    return ws


def _tarhiyat_dolgu(vurgu):
    if vurgu == "vurgu1":
        return DOLGU_SARI
    if vurgu == "vurgu2":
        return DOLGU_YESIL
    return None


def _yil_uyum_sayfasi(wb, sonuc):
    """Cok yilli incelemelerde yil bazinda beyan uyumu."""
    ws = wb.create_sheet("Yıl Uyumu")
    basliklar = [("Yıl", 8), ("Dönem", 8), ("Tespitli dönem", 14), ("Beyan matrah", 16),
                 ("Olması gereken matrah", 20), ("Matrah farkı", 16), ("Beyan ödenecek", 16),
                 ("Olması gereken ödenecek", 21), ("Re'sen tarhı gereken", 19),
                 ("Tespit etkisi", 15), ("Devir etkisi", 15), ("Beyan aritmetik farkı", 20),
                 ("Yıl sonu devir (beyan)", 20), ("Yıl sonu devir (olması gereken)", 26)]
    for i, (_ad, genislik) in enumerate(basliklar, start=1):
        ws.column_dimensions[get_column_letter(i)].width = genislik
    for i, (ad, _g) in enumerate(basliklar, start=1):
        _yaz(ws, 1, i, ad, FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
    kodlar = ["yil", "donem_sayisi", "tespit_donem_sayisi", "matrah_beyan",
              "matrah_olmasi_gereken", "matrah_farki", "odenecek_beyan",
              "odenecek_olmasi_gereken", "resen_tarhi_gereken", "tespit_etkisi",
              "devir_etkisi", "beyan_hatasi_etkisi", "devir_cikis_beyan", "devir_cikis"]
    tamsayi = {"yil", "donem_sayisi", "tespit_donem_sayisi"}
    satir = 2
    for s in sonuc.get("yil_uyumu") or []:
        for i, kod in enumerate(kodlar, start=1):
            sayim = kod in tamsayi
            _yaz(ws, satir, i, s[kod], FONT, ORTA if sayim else SAG,
                 bicim="0" if sayim else SAYI_BICIMI)
        satir += 1
    satir += 1
    _not_satiri(ws, satir,
                "Bu sayfa, seri bir kez hesaplanarak üretilen bir çıktıdır; "
                "formül içermez. Yıl sayfalarında bir değişiklik yaparsanız "
                "burayı uygulamadan yeniden dışa aktarınız.", len(basliklar))
    return ws


def _etki_sayfasi(wb, sonuc):
    """Her tespitin donem donem ayri ayri katkisi."""
    analiz = sonuc.get("kaynak_analizi") or {}
    kaynaklar = analiz.get("kaynaklar") or []
    if not kaynaklar:
        return None
    ws = wb.create_sheet("Tespit Etkisi")
    ws.column_dimensions["A"].width = 16
    for i in range(2, len(kaynaklar) + 4):
        ws.column_dimensions[get_column_letter(i)].width = 17

    satir = 1
    _yaz(ws, satir, 1, "TESPİTLERİN KAYNAK BAZINDA ETKİSİ", FONT_BASLIK, SOL,
         DOLGU_BASLIK, bicim=None)
    ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=len(kaynaklar) + 3)
    satir += 2

    _yaz(ws, satir, 1, "Tespit dönemi", FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
    for i, ad in enumerate(["Matrah ilavesi", "Hesaplanan KDV ilavesi", "İndirim reddi",
                            "Devir çıkarması", "Yüklenilen reddi", "Etkilenen dönem",
                            "Seri geneli ödenecek etkisi"], start=2):
        _yaz(ws, satir, i, ad, FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
    satir += 1
    for k in kaynaklar:
        _yaz(ws, satir, 1, k["etiket"], FONT_BOLD, SOL, DOLGU_ELESTIRI, bicim=None)
        for i, kod in enumerate(["matrah_ilave", "hesaplanan_kdv_ilave", "indirim_cikar",
                                 "devir_cikar", "yuklenilen_cikar"], start=2):
            _yaz(ws, satir, i, k[kod], FONT, SAG, DOLGU_ELESTIRI)
        _yaz(ws, satir, 7, k["etkilenen_donem_sayisi"], FONT, ORTA, DOLGU_ELESTIRI, bicim=None)
        _yaz(ws, satir, 8, k["toplam_odenecek_etkisi"], FONT_BOLD, SAG, DOLGU_FARK)
        satir += 1
    satir += 1

    etkilesim_var = analiz.get("etkilesim_var")
    for alan, baslik in (("odenecek", "ÖDENMESİ GEREKEN KDV'YE KATKI"),
                         ("sonraki_devir", "SONRAKİ DÖNEME DEVREDEN KDV'YE KATKI")):
        satir = _baslik_satiri(ws, satir, baslik, len(kaynaklar) + 3)
        _yaz(ws, satir, 1, "Dönem", FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
        for i, k in enumerate(kaynaklar, start=2):
            _yaz(ws, satir, i, k["etiket"], FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
        sutun = len(kaynaklar) + 2
        if etkilesim_var:
            _yaz(ws, satir, sutun, "Etkileşim", FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
            sutun += 1
        _yaz(ws, satir, sutun, "Toplam", FONT_BOLD, ORTA, DOLGU_TOPLAM, bicim=None)
        satir += 1
        for d in analiz.get("donemler") or []:
            if abs(d["toplam"].get(alan, 0)) < 0.005 and not any(
                    abs(d["paylar"][k["anahtar"]].get(alan, 0)) > 0.005 for k in kaynaklar):
                continue
            _yaz(ws, satir, 1, d["etiket"], FONT, SOL, bicim=None)
            for i, k in enumerate(kaynaklar, start=2):
                _yaz(ws, satir, i, d["paylar"][k["anahtar"]][alan], FONT, SAG)
            sutun = len(kaynaklar) + 2
            if etkilesim_var:
                _yaz(ws, satir, sutun, d["etkilesim"][alan], FONT, SAG)
                sutun += 1
            _yaz(ws, satir, sutun, d["toplam"][alan], FONT_BOLD, SAG, DOLGU_TOPLAM)
            satir += 1
        satir += 1

    if etkilesim_var:
        _yaz(ws, satir, 1,
             "Etkileşim: KDV hesabı doğrusal olmadığından (ödenecek/devreden eşiğinde "
             "alt-üst sınır) tespitlerin tek tek etkilerinin toplamı, hepsi birlikte "
             "uygulandığındaki farkı tam vermeyebilir. Aradaki bakiye kaynaklara "
             "dağıtılmaz, ayrıca gösterilir.", FONT, SOL, DOLGU_FARK, bicim=None)
        ws.merge_cells(start_row=satir, start_column=1, end_row=satir,
                       end_column=len(kaynaklar) + 3)
        satir += 2
    _not_satiri(ws, satir,
                "Bu sayfa, her tespit tek tek açılıp seri yeniden hesaplanarak "
                "üretilir; formül içermez. Yıl sayfalarında bir değişiklik "
                "yaparsanız burayı uygulamadan yeniden dışa aktarınız.",
                len(kaynaklar) + 3)
    return ws


def _duzeltme_sayfasi(wb, duzeltme_bloklari):
    """Rapora yapistirilan duzeltme beyannameleri tablosu.

    Elde kullanilan tablonun birebir karsiligidir: yalnizca duzeltme
    beyannameleri, kendi rakamlariyla ve duzeltme gerekcesiyle. Basligi
    "Dönemi <yil>" oldugu icin her yil ayri bir tablodur.
    """
    from .beyannameler import DUZELTME_KOLONLARI

    if not duzeltme_bloklari:
        return None
    ws = wb.create_sheet("Düzeltme Beyannameleri")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    for i in range(3, len(DUZELTME_KOLONLARI)):
        ws.column_dimensions[get_column_letter(i)].width = 17
    ws.column_dimensions[get_column_letter(len(DUZELTME_KOLONLARI))].width = 46

    metin_kodlari = {"donem", "tarih", "gerekce"}
    satir = 1
    for blok in duzeltme_bloklari:
        _yaz(ws, satir, 1, "MÜKELLEFİN VERDİĞİ DÜZELTME BEYANNAMELERİ",
             FONT_BASLIK, SOL, DOLGU_BASLIK, bicim=None)
        ws.merge_cells(start_row=satir, start_column=1, end_row=satir,
                       end_column=len(DUZELTME_KOLONLARI))
        satir += 1

        for i, (kod, etiket) in enumerate(DUZELTME_KOLONLARI):
            baslik = "Dönemi\n%s" % blok["yil"] if kod == "donem" else etiket
            h = _yaz(ws, satir, i + 1, baslik, FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
            h.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        ws.row_dimensions[satir].height = 30
        satir += 1

        for kayit in blok["satirlar"]:
            for i, (kod, _etiket) in enumerate(DUZELTME_KOLONLARI):
                deger = kayit.get(kod)
                if kod in metin_kodlari:
                    hiza = SOL if kod == "gerekce" else ORTA
                    h = _yaz(ws, satir, i + 1, deger or "", FONT, hiza, bicim=None)
                    if kod == "gerekce":
                        h.alignment = SOL
                else:
                    _yaz(ws, satir, i + 1, float(deger or 0.0), FONT, SAG)
            ws.row_dimensions[satir].height = 30
            satir += 1
        satir += 1
    return ws


FATURA_KOLONLARI = [
    ("satici_vkn", "Satıcı VKN", 15, "metin"),
    ("satici_unvan", "Satıcı Unvanı", 32, "metin"),
    ("kullanma", "Kullanma Durumu", 18, "metin"),
    ("fatura_no", "Fatura No", 22, "metin"),
    ("tarih", "Fatura Tarihi", 14, "metin"),
    ("kayit_donemi", "Kayıt Dönemi", 13, "metin"),
    ("matrah", "Matrah", 16, "tutar"),
    ("kdv", "KDV", 15, "tutar"),
    ("toplam", "Genel Toplam", 16, "tutar"),
    ("mal_cinsi", "Malın Cinsi", 20, "metin"),
    ("yevmiye", "Yevmiye Tarih / No", 20, "metin"),
    ("kaynak", "Kaynak Dosya", 34, "metin"),
]


def _fatura_sayfasi(wb, fatura_bloku):
    """Sahte belge kullanmaya konu fatura dokumu.

    Yalnizca tarhiyata dahil edilen alis faturalari yazilir; boylece sayfanin
    toplami, tespitlerdeki indirim reddi ve rapordaki fatura tablosuyla ayni
    kumeye dayanir.
    """
    if not fatura_bloku or not fatura_bloku.get("faturalar"):
        return None
    ws = wb.create_sheet("Faturalar")
    for i, (_kod, _etiket, genislik, _tur) in enumerate(FATURA_KOLONLARI, 1):
        ws.column_dimensions[get_column_letter(i)].width = genislik

    _yaz(ws, 1, 1, "SAHTE BELGE KULLANMAYA KONU FATURALAR",
         FONT_BASLIK, SOL, DOLGU_BASLIK, bicim=None)
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(FATURA_KOLONLARI))
    for i, (_kod, etiket, _g, _t) in enumerate(FATURA_KOLONLARI, 1):
        _yaz(ws, 2, i, etiket, FONT_BOLD, ORTA, DOLGU_BASLIK, bicim=None)
    ws.row_dimensions[2].height = 28

    satir = 3
    for f in fatura_bloku["faturalar"]:
        for i, (kod, _e, _g, tur) in enumerate(FATURA_KOLONLARI, 1):
            deger = f.get(kod)
            if tur == "tutar":
                _yaz(ws, satir, i, float(deger or 0.0), FONT, SAG)
            else:
                _yaz(ws, satir, i, deger or "", FONT,
                     SOL if kod in ("satici_unvan", "mal_cinsi", "kaynak") else ORTA,
                     bicim=None)
        satir += 1

    _yaz(ws, satir, 1, "TOPLAM", FONT_BOLD, SOL, DOLGU_TOPLAM, bicim=None)
    for i, (kod, _e, _g, tur) in enumerate(FATURA_KOLONLARI, 1):
        if i == 1:
            continue
        deger = fatura_bloku["toplam"].get(kod) if tur == "tutar" else ""
        _yaz(ws, satir, i, float(deger or 0.0) if tur == "tutar" else "",
             FONT_BOLD, SAG if tur == "tutar" else ORTA,
             DOLGU_TOPLAM, bicim=SAYI_BICIMI if tur == "tutar" else None)

    satir += 2
    _yaz(ws, satir, 1, "Tabloda yalnızca tarhiyata dahil edilen alış faturaları "
         "yer alır; dahil işareti kaldırılan satırlar yazılmaz.",
         FONT_NOT, SOL, bicim=None)
    ws.merge_cells(start_row=satir, start_column=1, end_row=satir,
                   end_column=len(FATURA_KOLONLARI))
    return ws


def calisma_olustur(dosya_yolu, inceleme, yillar, sonuc, bulgular=None,
                    duzeltme_bloklari=None, fatura_bloku=None, ziya_kati=1):
    """Excel calisma dosyasini uretir ve yola yazar."""
    wb = Workbook()
    wb.remove(wb.active)
    duzen = None
    onceki = None
    for yil_kaydi in sorted(yillar, key=lambda y: y["yil"]):
        duzen = _yil_sayfasi(wb, yil_kaydi, sonuc["donemler"], onceki)
        onceki = (duzen["yil"], duzen["ay_sayisi"])
    _duzeltme_sayfasi(wb, duzeltme_bloklari or [])
    _fatura_sayfasi(wb, fatura_bloku)
    _etki_sayfasi(wb, sonuc)
    _yil_uyum_sayfasi(wb, sonuc)
    if duzen:
        # Satir duzeni butun yil sayfalarinda aynidir; sonuncusu yeterlidir
        _tarhiyat_sayfasi(wb, inceleme, sonuc, duzen, ziya_kati)
        _ozet_sayfasi(wb, inceleme, sonuc, bulgular or [], duzen)
    os.makedirs(os.path.dirname(dosya_yolu), exist_ok=True)
    wb.save(dosya_yolu)
    return dosya_yolu
