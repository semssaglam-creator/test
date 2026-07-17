"""Sorgu sonuclarini iki sayfali Excel dosyasina aktarir.

Sayfa 1: Fatura Ozeti  (fatura basina bir satir)
Sayfa 2: Kalem Detayi  (mal/hizmet satiri basina bir satir)
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .ayristirici import iso_to_tr

FONT_BASLIK = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="Calibri", size=11)
DOLGU_BASLIK = PatternFill("solid", fgColor="1F3A5F")
INCE = Side(style="thin", color="B7C1CC")
KENARLIK = Border(left=INCE, right=INCE, top=INCE, bottom=INCE)
HIZA_SOL = Alignment(horizontal="left", vertical="center")
HIZA_SAG = Alignment(horizontal="right", vertical="center")
PARA_BICIM = "#,##0.00"

YON_ADLARI = {"alis": "Alış", "satis": "Satış", "": ""}
DURUM_ADLARI = {"tamam": "Tamam", "kontrol": "Kontrol gerekli",
                "ocr_bekliyor": "OCR bekliyor"}
KAYNAK_ADLARI = {"metin": "PDF metni", "ocr": "OCR", "elle": "Elle giriş"}

OZET_KOLONLAR = [
    ("Fatura No", 20), ("Tarih", 12), ("Yön", 8),
    ("Satıcı Ünvanı", 34), ("Satıcı VKN/TCKN", 16),
    ("Alıcı Ünvanı", 34), ("Alıcı VKN/TCKN", 16),
    ("Mal/Hizmet Toplamı", 17), ("İskonto", 12), ("KDV", 13),
    ("Vergiler Dahil Toplam", 17), ("Ödenecek Tutar", 15),
    ("Kaynak", 11), ("Durum", 15), ("Uyarılar", 40), ("Dosya", 28),
]

KALEM_KOLONLAR = [
    ("Fatura No", 20), ("Tarih", 12), ("Yön", 8),
    ("Satıcı Ünvanı", 30), ("Alıcı Ünvanı", 30),
    ("Sıra", 6), ("Mal / Hizmet", 44), ("Miktar", 11), ("Birim", 9),
    ("Birim Fiyat", 12), ("KDV %", 8), ("KDV Tutarı", 12), ("Tutar", 14),
]


def _sayfa_hazirla(ws, kolonlar):
    for i, (ad, genislik) in enumerate(kolonlar, 1):
        h = ws.cell(row=1, column=i, value=ad)
        h.font = FONT_BASLIK
        h.fill = DOLGU_BASLIK
        h.border = KENARLIK
        h.alignment = HIZA_SOL
        ws.column_dimensions[get_column_letter(i)].width = genislik
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(kolonlar))}1"


def _satir_yaz(ws, satir_no, degerler, para_kolonlari):
    for i, deger in enumerate(degerler, 1):
        h = ws.cell(row=satir_no, column=i, value=deger)
        h.font = FONT_NORMAL
        h.border = KENARLIK
        if i in para_kolonlari:
            h.alignment = HIZA_SAG
            if isinstance(deger, (int, float)):
                h.number_format = PARA_BICIM
        else:
            h.alignment = HIZA_SOL


def rapor_uret(basliklar, kalemler):
    """Fatura basliklari + kalem satirlarindan xlsx byte'lari uretir."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Fatura Özeti"
    _sayfa_hazirla(ws, OZET_KOLONLAR)
    para_kolonlari = {8, 9, 10, 11, 12}
    for satir_no, f in enumerate(basliklar, 2):
        _satir_yaz(ws, satir_no, [
            f.get("fatura_no", ""), iso_to_tr(f.get("fatura_tarihi", "")),
            YON_ADLARI.get(f.get("yon", ""), f.get("yon", "")),
            f.get("satici_unvan", ""), f.get("satici_vkn", ""),
            f.get("alici_unvan", ""), f.get("alici_vkn", ""),
            f.get("mal_hizmet_toplam"), f.get("toplam_iskonto"),
            f.get("kdv_toplam"), f.get("vergiler_dahil_toplam"),
            f.get("odenecek_tutar"),
            KAYNAK_ADLARI.get(f.get("kaynak", ""), f.get("kaynak", "")),
            DURUM_ADLARI.get(f.get("durum", ""), f.get("durum", "")),
            f.get("uyarilar", ""), f.get("dosya_adi", ""),
        ], para_kolonlari)

    ws2 = wb.create_sheet("Kalem Detayı")
    _sayfa_hazirla(ws2, KALEM_KOLONLAR)
    para_kolonlari = {8, 10, 11, 12, 13}
    for satir_no, k in enumerate(kalemler, 2):
        _satir_yaz(ws2, satir_no, [
            k.get("fatura_no", ""), iso_to_tr(k.get("fatura_tarihi", "")),
            YON_ADLARI.get(k.get("yon", ""), k.get("yon", "")),
            k.get("satici_unvan", ""), k.get("alici_unvan", ""),
            k.get("sira"), k.get("mal_hizmet", ""), k.get("miktar"),
            k.get("birim", ""), k.get("birim_fiyat"), k.get("kdv_orani"),
            k.get("kdv_tutari"), k.get("tutar"),
        ], para_kolonlari)

    tampon = io.BytesIO()
    wb.save(tampon)
    return tampon.getvalue()
