"""Duzeltme fisi Excel ciktisi.

Vergi dairesinin duzenledigi "duzeltme fisi" formunun karsiligidir. Bir yil
icin ayni sayfada iki tablo alt alta gelir:

  - MUKELLEF TARAFINDAN SURESINDE BEYAN EDILEN KDV DOKUMLERI
      Beyan Verisi sekmesindeki rakamlar (hesap._beyan_ozeti).
  - MUKELLEF ADINA OLMASI GEREKEN KDV BEYANNAME DOKUMLERI
      Inceleme tespitleri uygulandiktan ve devir zinciri yeniden kurulduktan
      sonraki rakamlar (hesap._donem_hesapla).

Ikinci tablonun altina iki satir eklenir:

  TARHI GEREKEN VERGI
      Tarhiyat Ozeti sekmesindeki "Re'sen Tarhi Gereken KDV" sutunuyla ayni
      tutardir (hesap._tarhiyat_satiri -> "resen_tarhi_gereken"): olmasi
      gereken odenecek KDV ile beyan edilenin farkinin yalnizca mukellef
      aleyhine olan yonu.
  SONRAKI DON DEV KDV UYUMSUZLUK TUTARI
      Beyan edilen sonraki doneme devreden KDV ile olmasi gerekenin farki;
      fazla ve yersiz devredilen tutardir. Formun altindaki gerekce metni bu
      tutara dayanir.

Form yil basligi tasidigi icin her yil AYRI bir sayfaya yazilir.

Incelemenin kapsamadigi aylar bos birakilir; sifir yazilmaz. Sifir yazmak, o
ay icin beyan yokmus ya da tum tutarlar sifirmis izlenimi verir ve fis
imzalanan bir belge oldugu icin bu ayrim onemlidir.
"""
import os
import sys
from datetime import date, datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LIB_DIR = os.path.join(BASE_DIR, "lib")
if LIB_DIR not in sys.path:
    sys.path.insert(0, LIB_DIR)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Formdaki ay basliklari. satirlar.AYLAR_BUYUK yerine burada ayrica tutulur:
# Python'un upper() cagrisi "Nisan" -> "NISAN" uretir, formda ise noktali
# "NİSAN" yazar. Fis imzalanan bir belge oldugu icin basliklar birebir korunur.
AYLAR_FIS = ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN",
             "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]

# Fis tablosunun satirlari: (etiket, hesap motorundaki alan adi).
# Ayni alan adlari hem beyan ozetinde hem elestirili sonucta bulunur; bu yuzden
# iki tablo tek listeden uretilir.
FIS_SATIRLARI = [
    ("MATRAH TOPLAMI", "matrah"),
    ("HESAPLANAN KDV", "hesaplanan"),
    ("İLAVE EDİLECEK KDV", "ilave_edilecek"),
    ("TOPLAM HESAPLANAN KDV", "toplam_kdv"),
    ("ÖNCEKİ DÖN. DEVR. İND. KDV", "onceki_devir"),
    ("BU DÖN. AİT İND. KDV", "bu_donem_indirim"),
    ("103+104+105 TOP.", "diger_indirim"),
    ("İNDİRİMLER TOPLAMI", "indirimler"),
    ("İADE EDİLMESİ GEREKEN KDV", "iade"),
    ("ÖDENMESİ GEREKEN KDV", "odenecek"),
    ("SONRAKİ DÖN. DEVREDEN KDV", "sonraki_devir"),
]

# Fisin altindaki gerekce paragrafi, devir tablosunun neye dayandigini soyler.
# Bu yuzden duzenlenme nedeni ile paragraftaki dayanak birlikte degismelidir;
# ikisi asagida cift halinde tutulur. Arayuz nedeni bu listeden onerir,
# kullanici baska bir neden de yazabilir.
ACIKLAMA_KALIBI = (
    "Adı geçen mükellefin %s sonraki vergilendirme dönemi KDV beyannamelerinin "
    "fazla ve yersiz olarak devreden KDV yönünden re’sen düzeltilmesi "
    "neticesinde yukarıdaki tablodaki vergilendirme dönemlerinde 213 sayılı "
    "V.U.K.’nun düzeltmeye ilişkin 116. ve müteakip maddeleri gereğince "
    "KDV’nin tarh edilmesi gerekmektedir."
)

NEDEN_SECENEKLERI = [
    ("VERGİ İNCELEME RAPORUNA İSTİNADEN",
     "ekli vergi inceleme raporu nedeniyle"),
    ("TAKDİR KOMİSYONU KARARINA İSTİNADEN",
     "ekli takdir komisyonu kararları nedeniyle"),
    ("DEVREDEN KDV BEYAN UYUMSUZLUĞU",
     "devreden KDV tutarlarında beyanlar arasında tespit edilen uyumsuzluk "
     "nedeniyle"),
]

VARSAYILAN_NEDEN = NEDEN_SECENEKLERI[0][0]
VARSAYILAN_ACIKLAMA = ACIKLAMA_KALIBI % NEDEN_SECENEKLERI[0][1]
VARSAYILAN_IMZALAR = [
    {"ad": "", "unvan": "GELİR UZMANI"},
    {"ad": "", "unvan": "MÜDÜR YRD."},
    {"ad": "", "unvan": "VERGİ DAİRESİ MÜDÜRÜ"},
]

# ------------------------------------------------------------------- bicimler
# Fisin tamami Times New Roman 10 punto. Baslik, bant ve etiketler boyutla
# degil kalinlikla ayrilir; boylece butun sayfa tek punto tutar.
AD = "Times New Roman"
PUNTO = 10
FONT_ETIKET = Font(name=AD, size=PUNTO, bold=True)
FONT_DEGER = Font(name=AD, size=PUNTO)
FONT_KUNYE = Font(name=AD, size=PUNTO, bold=True)
FONT_KUNYE_DEGER = Font(name=AD, size=PUNTO)
FONT_YIL = Font(name=AD, size=PUNTO, bold=True)
FONT_FIS = Font(name=AD, size=PUNTO, bold=True)
FONT_BANT = Font(name=AD, size=PUNTO, bold=True)
FONT_ACIKLAMA = Font(name=AD, size=PUNTO)
FONT_IMZA = Font(name=AD, size=PUNTO)

SARI = PatternFill("solid", fgColor="FFFF00")
KIRMIZI = PatternFill("solid", fgColor="FF6D6D")
ACIK_GRI = PatternFill("solid", fgColor="D9D9D9")

INCE = Side(style="hair", color="000000")
KALIN = Side(style="double", color="000000")

SOL = Alignment(horizontal="left", vertical="center", wrap_text=True)
SOL_DAR = Alignment(horizontal="left", vertical="center")
ORTA = Alignment(horizontal="center", vertical="center", wrap_text=True)
SAG = Alignment(horizontal="right", vertical="center")
SAG_UST = Alignment(horizontal="right", vertical="top")

# Eksi tutarlar isaretiyle gosterilir. Formun kendisinde bazi satirlarda
# isaretsiz bir bicim duruyor; uyumsuzluk satiri eksiye donebildigi icin
# (mukellef devri oldugundan az beyan etmisse) isaretin gorunmesi sarttir.
SAYI_BICIMI = "#,##0.00;\\-#,##0.00"

ILK_SUTUN = 1    # A: etiket sutunu
SON_SUTUN = 13   # M: ARALIK

# Imza bloklari: B:E, F:I, J:M. Ucu de dorder sutun, yani birbirine esit
# genislikte. Etiket sutunu A disarida birakilir; o sutun otekilerden genis
# oldugu icin iceri alinsaydi bloklardan biri digerlerinden genis kalirdi.
IMZA_BLOKLARI = [(2, 5), (6, 9), (10, 13)]


def _kenar(sutun, ust_kalin=False, alt_kalin=False):
    """Tablo hucresinin cercevesi: disa bakan kenarlar cift, ic kenarlar ince."""
    return Border(
        left=KALIN if sutun == ILK_SUTUN else INCE,
        right=KALIN if sutun == SON_SUTUN else INCE,
        top=KALIN if ust_kalin else INCE,
        bottom=KALIN if alt_kalin else INCE,
    )


def _kunye_satiri(ws, satir, etiket, deger, ust_kalin=False, alt_kalin=False,
                  deger_font=None):
    """Ust bloktaki tek satirlik kunye alani (etiket + B:M birlesik deger)."""
    hucre = ws.cell(row=satir, column=1, value=etiket)
    hucre.font = FONT_KUNYE
    hucre.alignment = SOL
    hucre.border = Border(left=KALIN, right=INCE,
                          top=KALIN if ust_kalin else INCE,
                          bottom=KALIN if alt_kalin else INCE)
    ws.merge_cells(start_row=satir, start_column=2, end_row=satir, end_column=SON_SUTUN)
    deger_hucre = ws.cell(row=satir, column=2, value=deger)
    deger_hucre.font = deger_font or FONT_KUNYE_DEGER
    deger_hucre.alignment = SOL
    # Birlesik alanin sag kenari M sutununda; cerceve icin o hucre de bicimlenir
    for sutun in range(2, SON_SUTUN + 1):
        ws.cell(row=satir, column=sutun).border = Border(
            left=INCE, right=KALIN if sutun == SON_SUTUN else INCE,
            top=KALIN if ust_kalin else INCE,
            bottom=KALIN if alt_kalin else INCE)


def _bant(ws, satir, metin):
    """Sari zeminli tablo basligi (A:M birlesik)."""
    ws.merge_cells(start_row=satir, start_column=1, end_row=satir, end_column=SON_SUTUN)
    for sutun in range(1, SON_SUTUN + 1):
        hucre = ws.cell(row=satir, column=sutun)
        hucre.fill = SARI
        hucre.border = Border(left=KALIN if sutun == ILK_SUTUN else None,
                              right=KALIN if sutun == SON_SUTUN else None,
                              top=KALIN)
    hucre = ws.cell(row=satir, column=1, value=metin)
    hucre.font = FONT_BANT
    hucre.alignment = ORTA


def _ay_basligi(ws, satir):
    hucre = ws.cell(row=satir, column=1, value="DÖNEMİ")
    hucre.font = FONT_ETIKET
    hucre.alignment = ORTA
    hucre.border = _kenar(1)
    for i, ay in enumerate(AYLAR_FIS):
        hucre = ws.cell(row=satir, column=2 + i, value=ay)
        hucre.font = FONT_ETIKET
        hucre.alignment = ORTA
        hucre.border = _kenar(2 + i)


def _veri_satiri(ws, satir, etiket, degerler, ay_sayisi, dolgu=None,
                 alt_kalin=False, kalin=False, vurgu=None):
    """Etiket + 12 aylik tutar. ay_sayisi disindaki aylar bos birakilir.

    dolgu verilirse etiketiyle tutarlariyla satirin tamamina uygulanir;
    kalin ise tutarlar da etiket gibi kalin yazilir. vurgu verilirse
    yalnizca SIFIRDAN FARKLI tutar tasiyan aylara uygulanir; boylece hangi
    donemde tarhiyat ciktigi tabloya bakar bakmaz gorulur.
    """
    hucre = ws.cell(row=satir, column=1, value=etiket)
    hucre.font = FONT_ETIKET
    hucre.alignment = SOL
    hucre.border = _kenar(1, alt_kalin=alt_kalin)
    if dolgu is not None:
        hucre.fill = dolgu
    for ay in range(12):
        hucre = ws.cell(row=satir, column=2 + ay)
        if ay < ay_sayisi:
            hucre.value = degerler[ay]
        hucre.font = FONT_ETIKET if kalin else FONT_DEGER
        hucre.alignment = SAG
        hucre.number_format = SAYI_BICIMI
        hucre.border = _kenar(2 + ay, alt_kalin=alt_kalin)
        if vurgu is not None and ay < ay_sayisi and abs(degerler[ay]) > 0.005:
            hucre.fill = vurgu
        elif dolgu is not None:
            hucre.fill = dolgu


def _imzalari_coz(ham):
    """Imza girdisini (ad soyad, unvan) uclusune cevirir; her zaman 3 oge.

    Yalnizca metin verilmisse unvan sayilir: alan once yalnizca unvan tasiyordu,
    imza sahibinin adi sonradan eklendi ve o bicimde kaydedilmis calismalarin
    yeniden acilabilmesi gerekir.
    """
    cozum = []
    for oge in list(ham or [])[:3]:
        if isinstance(oge, dict):
            cozum.append(((oge.get("ad") or "").strip(),
                          (oge.get("unvan") or "").strip()))
        else:
            cozum.append(("", str(oge or "").strip()))
    while len(cozum) < 3:
        cozum.append(("", ""))
    return cozum


def _tarihe_cevir(ham):
    """'02.09.2026' / '2026-09-02' metnini tarihe cevirir; olmazsa None."""
    if isinstance(ham, (datetime, date)):
        return ham
    metin = (ham or "").strip()
    if not metin:
        return None
    for bicim in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(metin, bicim).date()
        except ValueError:
            continue
    return None


def _yil_sayfasi(wb, yil, donemler, ay_sayisi, inceleme, fis):
    """Bir yilin fis sayfasini yazar.

    donemler: o yila ait, kronolojik sirali hesap sonuclari
    ay_sayisi: incelemenin o yilda kapsadigi ay sayisi
    """
    ws = wb.create_sheet(title=("%s Düzeltme Fişi" % yil)[:31])
    ws.sheet_view.showGridLines = False

    # Genislikler 10 puntoya gore secildi. Ay sutunu "123.456.789,00"
    # genisligindedir: daha dar yapilirsa buyuk tutarlar ##### olarak basilir.
    #
    # Etiket sutunu, beyanname satir adlarinin cogu tek satira sigacak
    # genisliktedir. Bunun ikinci bir islevi var: tablonun toplam eni,
    # sigdirmada BOY degil EN sinirlayici olacak kadar buyur. Boy sinirlayici
    # olsaydi Excel sayfayi boya gore kucultur, en tam dolmaz ve sayfanin
    # saginda solunda bosluk kalirdi.
    ws.column_dimensions["A"].width = 32.0
    for sutun in range(2, SON_SUTUN + 1):
        ws.column_dimensions[get_column_letter(sutun)].width = 12.5

    def al(kaynak, alan):
        return [d[kaynak][alan] for d in donemler]

    # --------------------------------------------------------------- kunye
    _kunye_satiri(ws, 1, "DÖNEMİ", yil, ust_kalin=True, deger_font=FONT_YIL)
    _kunye_satiri(ws, 2, "VERGİ KİMLİK NO", inceleme.get("vkn_tckn") or "")
    _kunye_satiri(ws, 3, "ÜNVANI", inceleme.get("ad_unvan") or "")
    _kunye_satiri(ws, 4, "DEVİR TABLOSUNUN DÜZENLENME NEDENİ",
                  fis.get("neden") or VARSAYILAN_NEDEN, alt_kalin=True)

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=11)
    baslik = ws.cell(row=5, column=1, value="DÜZELTME FİŞİ")
    baslik.font = FONT_FIS
    baslik.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=5, start_column=12, end_row=5, end_column=SON_SUTUN)
    cilt = ws.cell(row=5, column=12,
                   value="CİLT/SIRA NO: %s" % (fis.get("cilt_sira") or ""))
    cilt.font = FONT_FIS
    cilt.alignment = SOL_DAR
    # Baslik satiri ince cerceveli; iki birlesik alanin (A:K ve L:M) disa bakan
    # kenarlari cizilir, aradaki hucreler yalnizca ust/alt kenari tasir.
    for sutun in range(1, SON_SUTUN + 1):
        ws.cell(row=5, column=sutun).border = Border(
            left=INCE if sutun in (1, 12) else None,
            right=INCE if sutun in (11, 12, 13) else None,
            top=INCE, bottom=INCE)

    # -------------------------------------------------- beyan edilen tablosu
    _bant(ws, 6, "MÜKELLEF TARAFINDAN SÜRESİNDE BEYAN EDİLEN KDV "
                 "DÖKÜMLERİ  ( %s )" % yil)
    _ay_basligi(ws, 7)
    for sira, (etiket, alan) in enumerate(FIS_SATIRLARI):
        _veri_satiri(ws, 8 + sira, etiket, al("beyan", alan), ay_sayisi,
                     alt_kalin=(sira == len(FIS_SATIRLARI) - 1))

    # ------------------------------------------------ olmasi gereken tablosu
    _bant(ws, 20, "MÜKELLEF ADINA OLMASI GEREKEN KDV BEYANNAME "
                  "DÖKÜMLERİ  ( %s )" % yil)
    _ay_basligi(ws, 21)
    for sira, (etiket, alan) in enumerate(FIS_SATIRLARI):
        _veri_satiri(ws, 22 + sira, etiket, al("elestirili", alan), ay_sayisi)

    # Tarhiyat Ozeti'ndeki "Re'sen Tarhi Gereken KDV" sutununun aynisi
    # Fisin sonucunu tasiyan satir: tamami kalin ve acik gri zeminli; tarhiyat
    # cikan aylar ayrica kirmizi zeminle isaretlenir.
    _veri_satiri(ws, 33, "TARHI GEREKEN VERGİ",
                 [d["tarhiyat"]["resen_tarhi_gereken"] for d in donemler],
                 ay_sayisi, dolgu=ACIK_GRI, kalin=True, vurgu=KIRMIZI)
    # Fazla ve yersiz devredilen KDV: beyan edilen devir - olmasi gereken devir.
    # fark["sonraki_devir"] ters yonde (elestirili - beyan) tutuldugu icin
    # isareti cevrilir; boylece fazla devir pozitif cikar.
    _veri_satiri(ws, 34, "SONRAKİ DÖN DEV KDV UYUMSUZLUK TUTARI",
                 [round(-d["fark"]["sonraki_devir"], 2) + 0.0 for d in donemler],
                 ay_sayisi, alt_kalin=True)

    # ------------------------------------------------------ gerekce ve imza
    ws.merge_cells(start_row=36, start_column=1, end_row=36, end_column=SON_SUTUN)
    aciklama = ws.cell(row=36, column=1,
                       value=fis.get("aciklama") or VARSAYILAN_ACIKLAMA)
    aciklama.font = FONT_ACIKLAMA
    aciklama.alignment = SOL

    ws.merge_cells(start_row=37, start_column=1, end_row=37, end_column=SON_SUTUN)
    tarih_hucre = ws.cell(row=37, column=1)
    tarih = _tarihe_cevir(fis.get("tarih"))
    if tarih is None:
        tarih_hucre.value = (fis.get("tarih") or "").strip()
    else:
        tarih_hucre.value = tarih
        tarih_hucre.number_format = "dd/mm/yyyy"
    tarih_hucre.font = FONT_DEGER
    tarih_hucre.alignment = SAG_UST

    # Imza bloğu: ustte islak imza icin bosluk (37-38), sonra ad soyad ve
    # altinda unvan. Uc imza formdaki sutunlarda durur.
    imzalar = _imzalari_coz(fis.get("imzalar"))
    if not any(ad or unvan for ad, unvan in imzalar):
        imzalar = _imzalari_coz(VARSAYILAN_IMZALAR)
    orta = Alignment(horizontal="center", vertical="center")
    for (bas, son), (ad, unvan) in zip(IMZA_BLOKLARI, imzalar):
        for satir, metin in ((39, ad), (40, unvan)):
            hucre = ws.cell(row=satir, column=bas, value=metin)
            hucre.font = FONT_IMZA
            hucre.alignment = orta
            ws.merge_cells(start_row=satir, start_column=bas,
                           end_row=satir, end_column=son)

    # ----------------------------------------------------- satir yukseklikleri
    # Yukseklikler formun kendi oranlarini korur, ancak A4 yatay tek sayfaya
    # sigdirma olcegini yukseltmek icin bir miktar sikistirilmistir; 34. satir
    # etiketi iki satira tastigi icin daha yuksektir.
    # Yukseklik YALNIZCA birlesik hucre tasiyan satirlara ve bosluklara verilir.
    # Geri kalan satirlar yuksekliksiz birakilir; Excel onlari icerige gore
    # kendiliginden buyutur, boylece iki satira tasan bir etiket kirpilmaz.
    # Birlesik hucrelerde bu kendiliginden buyume calismadigi icin oradaki
    # yukseklikler elle verilmek zorundadir.
    yukseklikler = {5: 26.0, 6: 18.0, 19: 7.0, 20: 18.0,
                    35: 8.0, 36: 38.0, 37: 38.0, 38: 10.0, 39: 14.0, 40: 14.0}
    for satir, yukseklik in yukseklikler.items():
        ws.row_dimensions[satir].height = yukseklik

    # ------------------------------------------------------------- sayfa duzeni
    # A4 yatay, tek sayfaya tam sigar. fitToWidth/fitToHeight yalnizca
    # pageSetUpPr.fitToPage isaretliyken ve sabit bir olcek verilmemisken
    # gecerlidir; ikisi de 1 oldugu icin Excel sayfayi hem ene hem boya
    # sigdirir. Her yil ayri sayfada oldugundan her fis tek yaprak cikar.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.26
    ws.page_margins.bottom = 0.30
    # Ustbilgi/altbilgi payi varsayilan olarak 0.5 inch gelir ve ust/alt kenar
    # boslugundan buyuk oldugunda basilabilir yuksekligi o kadar kisar. Fis
    # ustbilgi tasimadigi icin bu pay kucultulur.
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_area = "A1:%s40" % get_column_letter(SON_SUTUN)
    return ws


def fis_olustur(dosya_yolu, inceleme, yillar, sonuc, fis=None):
    """Duzeltme fisi dosyasini uretir ve yolunu dondurur.

    inceleme : {"ad_unvan": ..., "vkn_tckn": ...}  (web_server._inceleme_bilgisi)
    yillar   : hesap motoruna verilen yil kayitlari (ay_sayisi buradan alinir)
    sonuc    : hesap.seri_hesapla ciktisi
    fis      : {"neden", "cilt_sira", "tarih", "aciklama", "imzalar"} - hepsi
               istege baglidir; verilmeyenler icin varsayilanlar kullanilir.
    """
    fis = fis or {}
    donemler = sonuc.get("donemler") or []
    if not donemler:
        raise ValueError("Hesaplanmis dönem yok.")
    ay_sayilari = {int(y["yil"]): int(y.get("ay_sayisi") or 12) for y in (yillar or [])}

    wb = Workbook()
    wb.remove(wb.active)
    for yil in sorted({d["yil"] for d in donemler}):
        alt = [d for d in donemler if d["yil"] == yil]
        alt.sort(key=lambda d: d["ay"])
        _yil_sayfasi(wb, yil, alt, ay_sayilari.get(yil, len(alt)), inceleme, fis)
    klasor = os.path.dirname(dosya_yolu)
    if klasor:
        os.makedirs(klasor, exist_ok=True)
    wb.save(dosya_yolu)
    return dosya_yolu
