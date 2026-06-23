"""Toplu tahakkuk Excel dosyalarini okur.

- .xls (eski Excel) icin vendored 'xlrd', .xlsx icin 'openpyxl' kullanilir.
- Baslik satiri otomatik bulunur (sutun adlarina gore).
- Bos satirlar bir ustteki mukellef/fis bilgisinin DEVAMI sayilir; ilgili
  alanlar yukaridan asagi tasinir (forward-fill). Vergi kodu ve odenecek tutar
  her satira ozeldir.
"""
import io

# Kolon adi -> i? anahtar eslesmesi icin normalize edilmis anahtar kelimeler.
_BASLIK_ESLES = [
    ("vergi_kimlik_no", ("vergi kimlik", "vkn", "kimlik no")),
    ("tahakkuk_fis_no", ("tahakkuk fis", "fis numara", "fis no")),
    ("islem_turu", ("islem tur",)),
    ("thk_turu", ("thk tur",)),
    ("vergi_donemi", ("vergi donem", "donem")),
    ("vergi_kodu", ("vergi kod",)),
    ("plaka_sasi", ("plaka", "sasi")),
    ("odenecek_tutar", ("odenecek", "tutar")),
]

# Bos satirda bir ustteki degerle doldurulacak (devam eden) alanlar.
TASINAN_ALANLAR = ("vergi_kimlik_no", "tahakkuk_fis_no", "islem_turu",
                   "thk_turu", "vergi_donemi", "plaka_sasi")

# Bu alanlardaki sayisal degerlerin bastaki sifirlari korunur (sola sifir doldurma).
SIFIR_DOLGU = {"vergi_kodu": 4, "islem_turu": 4, "thk_turu": 4, "vergi_donemi": 12}


class OkumaHatasi(Exception):
    """Kullaniciya gosterilecek okuma hatasi."""


def _tr_normalize(metin):
    metin = (metin or "").strip().lower()
    for a, b in (("ı", "i"), ("ğ", "g"), ("ü", "u"), ("ş", "s"),
                 ("ö", "o"), ("ç", "c"), ("İ", "i")):
        metin = metin.replace(a, b)
    return " ".join(metin.split())


def _hucre_metin(deger):
    """Hucre degerini temiz metne cevirir (tam sayilik float'larin .0'ini atar)."""
    if deger is None:
        return ""
    if isinstance(deger, float):
        if deger.is_integer():
            return str(int(deger))
        return repr(deger)
    return str(deger).strip()


def _tutar_coz(deger):
    if deger is None or deger == "":
        return 0.0
    if isinstance(deger, (int, float)):
        return float(deger)
    s = str(deger).strip()
    if not s:
        return 0.0
    # '1.234,56' (TR) ve '1234.56' (US) bicimlerini destekle.
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _satirlari_al(icerik, dosya_adi):
    """Excel icerigini ham hucre satirlari (liste-liste) olarak dondurur."""
    ad = (dosya_adi or "").lower()
    if ad.endswith(".xlsx") or ad.endswith(".xlsm"):
        return _xlsx_satirlar(icerik)
    if ad.endswith(".xls"):
        return _xls_satirlar(icerik)
    # Uzanti belirsizse once xls (eski bicim), sonra xlsx dene.
    try:
        return _xls_satirlar(icerik)
    except Exception:
        return _xlsx_satirlar(icerik)


def _xls_satirlar(icerik):
    import xlrd
    try:
        kitap = xlrd.open_workbook(file_contents=icerik)
    except Exception as exc:  # noqa: BLE001
        raise OkumaHatasi(f"Excel (.xls) dosyasi okunamadi: {exc}")
    sayfa = kitap.sheet_by_index(0)
    satirlar = []
    for r in range(sayfa.nrows):
        satirlar.append([sayfa.cell_value(r, c) for c in range(sayfa.ncols)])
    return satirlar


def _xlsx_satirlar(icerik):
    from openpyxl import load_workbook
    try:
        kitap = load_workbook(io.BytesIO(icerik), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise OkumaHatasi(f"Excel (.xlsx) dosyasi okunamadi: {exc}")
    sayfa = kitap.active
    return [list(satir) for satir in sayfa.iter_rows(values_only=True)]


def _baslik_satiri_bul(satirlar):
    """Sutun adlarini iceren satirin indeksini ve kolon->alan eslemesini bulur."""
    for idx, satir in enumerate(satirlar[:15]):
        normaller = [_tr_normalize(_hucre_metin(h)) for h in satir]
        eslesme = {}
        for kolon, metin in enumerate(normaller):
            if not metin:
                continue
            for alan, anahtarlar in _BASLIK_ESLES:
                if alan in eslesme:
                    continue
                if any(a in metin for a in anahtarlar):
                    eslesme[alan] = kolon
                    break
        # Zorunlu cekirdek alanlar bulunduysa bu satir basliktir.
        if {"vergi_kimlik_no", "tahakkuk_fis_no", "vergi_kodu", "odenecek_tutar"} <= set(eslesme):
            return idx, eslesme
    raise OkumaHatasi(
        "Baslik satiri bulunamadi. Dosyada 'Vergi Kimlik No', 'Tahakkuk Fiş Numarası', "
        "'Vergi Kodu' ve 'Ödenecek Olan Tutar' sutunlari olmali."
    )


def _kod_normalize(alan, metin):
    metin = (metin or "").strip()
    genislik = SIFIR_DOLGU.get(alan)
    if genislik and metin.isdigit() and len(metin) <= genislik:
        return metin.zfill(genislik)
    return metin


def kayitlardan_uret(ham_satirlar):
    """Ham alan->deger satirlarindan, bos alanlari forward-fill ederek kayit uretir.

    ham_satirlar: dict listesi; her dict alan adlarini (vergi_kimlik_no,
    tahakkuk_fis_no, ... odenecek_tutar) icerebilir. Bos/eksik alanlar bir
    ustteki mukellef/fis degeriyle doldurulur. Excel ve PDF okuyucular bunu
    ortak kullanir.
    """
    kayitlar = []
    tasinan = {alan: "" for alan in TASINAN_ALANLAR}

    for ham in ham_satirlar:
        deger = {}
        for alan, ham_deger in ham.items():
            if alan == "odenecek_tutar":
                deger[alan] = ham_deger
            else:
                deger[alan] = _kod_normalize(alan, _hucre_metin(ham_deger))

        # Tamamen bos satirlari atla.
        dolu = any(str(v).strip() for k, v in deger.items() if k != "odenecek_tutar")
        if not dolu and _tutar_coz(deger.get("odenecek_tutar")) == 0:
            continue

        # Devam eden (bos) alanlari ustteki degerle doldur.
        for alan in TASINAN_ALANLAR:
            if deger.get(alan):
                tasinan[alan] = deger[alan]
            else:
                deger[alan] = tasinan[alan]

        kayit = {
            "vergi_kimlik_no": deger.get("vergi_kimlik_no", ""),
            "tahakkuk_fis_no": deger.get("tahakkuk_fis_no", ""),
            "islem_turu": deger.get("islem_turu", ""),
            "thk_turu": deger.get("thk_turu", ""),
            "vergi_donemi": deger.get("vergi_donemi", ""),
            "vergi_kodu": deger.get("vergi_kodu", ""),
            "plaka_sasi": deger.get("plaka_sasi", ""),
            "odenecek_tutar": _tutar_coz(deger.get("odenecek_tutar")),
        }
        # Anlamsiz (ne vergi kodu ne tutar ne fis) satirlari atla.
        if not kayit["vergi_kodu"] and kayit["odenecek_tutar"] == 0 \
                and not kayit["tahakkuk_fis_no"]:
            continue
        kayitlar.append(kayit)
    return kayitlar


def excel_oku(icerik, dosya_adi=""):
    """Excel icerigini (bytes) okuyup kayit dict listesi dondurur.

    Bos satirlar bir ustteki mukellef/fis bilgisiyle doldurulur. Satir sayisi
    sinirsizdir.
    """
    satirlar = _satirlari_al(icerik, dosya_adi)
    if not satirlar:
        raise OkumaHatasi("Dosya bos görünüyor.")

    bas_idx, eslesme = _baslik_satiri_bul(satirlar)
    ham_satirlar = []
    for satir in satirlar[bas_idx + 1:]:
        ham = {}
        for alan, kolon in eslesme.items():
            ham[alan] = satir[kolon] if kolon < len(satir) else None
        ham_satirlar.append(ham)

    kayitlar = kayitlardan_uret(ham_satirlar)
    if not kayitlar:
        raise OkumaHatasi("Dosyada okunabilir veri satiri bulunamadi.")
    return kayitlar
