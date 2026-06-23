"""Sorgu sonucunu .xlsx olarak disa aktarir (vendored openpyxl)."""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .tarih_util import donem_metni

BASLIKLAR = [
    ("vergi_kimlik_no", "Vergi Kimlik No", 16),
    ("tahakkuk_fis_no", "Tahakkuk Fiş No", 24),
    ("vergi_donemi", "Vergi Dönemi", 18),
    ("vergi_kodu", "Vergi Kodu", 12),
    ("vergi_kodu_adi", "Vergi Kodu Adı", 40),
    ("odenecek_tutar", "Ödenecek Tutar", 16),
]


def sorgu_excel(kayitlar, kod_adlari=None):
    """Kayit listesinden xlsx bayt icerigi uretir."""
    kod_adlari = kod_adlari or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Tahakkuk Sorgusu"

    baslik_font = Font(bold=True, color="FFFFFF")
    baslik_dolgu = PatternFill("solid", fgColor="2F5496")
    for i, (_, baslik, genislik) in enumerate(BASLIKLAR, start=1):
        h = ws.cell(row=1, column=i, value=baslik)
        h.font = baslik_font
        h.fill = baslik_dolgu
        h.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = genislik

    toplam = 0.0
    for r, kayit in enumerate(kayitlar, start=2):
        tutar = kayit.get("odenecek_tutar") or 0
        toplam += tutar
        deger = {
            "vergi_kimlik_no": kayit.get("vergi_kimlik_no", ""),
            "tahakkuk_fis_no": kayit.get("tahakkuk_fis_no", ""),
            "vergi_donemi": donem_metni(kayit.get("vergi_donemi", "")),
            "vergi_kodu": kayit.get("vergi_kodu", ""),
            "vergi_kodu_adi": kod_adlari.get(kayit.get("vergi_kodu", ""), ""),
            "odenecek_tutar": tutar,
        }
        for i, (alan, _, _) in enumerate(BASLIKLAR, start=1):
            h = ws.cell(row=r, column=i, value=deger[alan])
            if alan == "odenecek_tutar":
                h.number_format = "#,##0.00"
                h.alignment = Alignment(horizontal="right")

    son = len(kayitlar) + 2
    ws.cell(row=son, column=4, value="TOPLAM").font = Font(bold=True)
    th = ws.cell(row=son, column=6, value=round(toplam, 2))
    th.font = Font(bold=True)
    th.number_format = "#,##0.00"
    ws.freeze_panes = "A2"

    tampon = io.BytesIO()
    wb.save(tampon)
    return tampon.getvalue()
