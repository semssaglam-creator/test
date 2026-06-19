"""Uzlasma tutanagi Excel ciktilarini uretir (3 sablon).

Sablonlar, GIB Dijital Vergi Dairesi tarafindan uretilen ornek tutanak
dosyalarinin (UZLASMA TUTANAGI / UZLASMA KOMISYON KARAR TUTANAGI) yapisina
gore hazirlanmistir.
"""
import os
import sys

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LIB_DIR = os.path.join(BASE_DIR, "lib")
if LIB_DIR not in sys.path:
    sys.path.insert(0, LIB_DIR)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .sayi_yaziya import tutar_yaziya

COL_WIDTHS = [23.71, 20.71, 15.71, 20.71, 15.71, 23.71]

FONT_NORMAL = Font(name="Times New Roman", size=11)
FONT_BOLD = Font(name="Times New Roman", size=11, bold=True)
FONT_TITLE = Font(name="Times New Roman", size=12, bold=True)

THIN = Side(style="thin")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_TITLE = Alignment(horizontal="center", vertical="center")
# Veri hucreleri: hucreye sigdir (tasma olmaz); paragraflar: iki yana yasli
SIGDIR = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
SIGDIR_SOL = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)
JUSTIFY = Alignment(horizontal="justify", vertical="center", wrap_text=True)
DOLGU_GRI = PatternFill("solid", fgColor="DDDDDD")

# Uzlasma Yonetmeligi Madde 13: uzlasmanin vaki olmamasi/temin edilememesi
# halinde dava acma ve surenin 15 gune uzamasi notu.
DAVA_SURESI_NOTU = (
    "NOT: Uzlaşma Yönetmeliğinin 13. maddesi uyarınca; mükellef, tarh edilen vergi "
    "ve kesilen cezaya karşı, tutanağın tebliğ tarihinden itibaren genel hükümler "
    "dairesinde ve yetkili vergi mahkemesi nezdinde dava açabilir. Dava açma süresi "
    "bitmiş veya 15 günden az kalmış ise bu süre, tutanağın tebliğ tarihinden "
    "itibaren 15 gün olarak uzar."
)


def _kur_kolon_genislikleri(ws):
    for i, genislik in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = genislik


def _sayfa_ayari(ws):
    """A4 dikey, genislige sigdir (orijinal GIB dosyasindaki ayarlar)."""
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


def _metin_satir_sayisi(metin, satir_genisligi=110):
    """Kac gorunur satira boluneceginin kaba tahmini (A:F birlesik hucre)."""
    toplam = 0
    for parca in str(metin).split("\n"):
        toplam += max(1, -(-len(parca) // satir_genisligi))
    return toplam


def _baslik_blogu(ws, kurum, tutanak_basligi):
    """1-7. satirlar: kurum bilgileri ve tutanak basligi (A1:F7 birlesik)."""
    satirlar = [
        "T.C.",
        "GELİR İDARESİ BAŞKANLIĞI",
        kurum.get("defterdarlik", ""),
        f"({kurum.get('vergi_dairesi', '')})",
        "",
        "",
        tutanak_basligi,
    ]
    for i, deger in enumerate(satirlar, start=1):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        c = ws.cell(row=i, column=1, value=deger)
        c.alignment = CENTER_TITLE
        c.font = FONT_TITLE if i == 7 else FONT_BOLD
    return 8  # sonraki bos satir


def _bilgi_satiri(ws, row, etiket, deger, sigdir=False):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    c1 = ws.cell(row=row, column=1, value=etiket)
    c1.font = FONT_BOLD
    c1.alignment = WRAP_LEFT
    c2 = ws.cell(row=row, column=3, value=deger)
    c2.font = FONT_NORMAL
    if sigdir:
        c2.alignment = SIGDIR_SOL
        ws.row_dimensions[row].height = 21
    else:
        c2.alignment = WRAP_LEFT
        # Uzun degerler (ornegin adres) sarildiginda satir yuksekligi yetsin
        satir_sayisi = _metin_satir_sayisi(deger, satir_genisligi=70)
        ws.row_dimensions[row].height = max(21, satir_sayisi * 15 + 6)
    return row + 1


def _ust_bilgi_blogu(ws, start_row, kurum, tutanak_no, tutanak_tarihi_metni, mukellef,
                      ara_baslik=""):
    row = start_row
    row = _bilgi_satiri(ws, row, "Tutanağın Tarihi / Davetiye T.Tarihi", tutanak_tarihi_metni)
    row = _bilgi_satiri(ws, row, "Tutanağın Sayısı", tutanak_no)
    if ara_baslik:
        row += 1  # buyuk harfli baslik ile ustteki satir arasinda bosluk
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=ara_baslik)
        c.font = FONT_BOLD
        c.alignment = CENTER_TITLE
        row += 1
    row = _bilgi_satiri(ws, row, "Mükellefin Adı Soyadı / Ünvanı", mukellef.get("ad_unvan", ""),
                         sigdir=True)
    row = _bilgi_satiri(ws, row, "Mükellefin Adresi", mukellef.get("adres", ""))
    row = _bilgi_satiri(ws, row, "Bağlı Bulunduğu Vergi Dairesi", kurum.get("vergi_dairesi", ""))
    row = _bilgi_satiri(ws, row, "Vergi Kimlik Numarası", mukellef.get("vkn_tckn", ""))
    return row


def _paragraf(ws, row, metin, yukseklik=None):
    # Paragraf basi girintisi: her paragrafin ilk satiri hafif sagdan baslar
    satirlar = []
    for parca in str(metin).split("\n"):
        if parca.strip() and not parca.startswith(" "):
            parca = "     " + parca
        satirlar.append(parca)
    metin = "\n".join(satirlar)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value=metin)
    c.alignment = JUSTIFY
    c.font = FONT_NORMAL
    if yukseklik is None:
        yukseklik = _metin_satir_sayisi(metin) * 15 + 10
    ws.row_dimensions[row].height = yukseklik
    return row + 1


def _tablo_kenarligi_tamamla(ws, r1, r2, c1=1, c2=6):
    """Birlesik hucrelerin ic hucrelerinde eksik kalan kenarliklari tamamlar."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER_ALL


def _baslik_satiri_boya(ws, r1, r2, c1=1, c2=6):
    """Tablo baslik hucrelerini hafif gri dolgu ile koyulastirir."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = DOLGU_GRI


def _hucre(ws, row, col, deger, font=None, align=None, border=True, num_format=None):
    c = ws.cell(row=row, column=col, value=deger)
    c.font = font or FONT_NORMAL
    c.alignment = align or CENTER
    if border:
        c.border = BORDER_ALL
    if num_format:
        c.number_format = num_format
    return c


def _unvan_normal(unvan):
    u = (unvan or "").strip().lower().replace("ş", "s").replace("ü", "u").replace("İ", "i")
    if u.startswith("ba"):
        return "baskan"
    if u.startswith("u"):
        return "uye"
    return u


def _imza_bloklari_4lu(ws, row, imzalayanlar, mukellef_adi):
    """Baskan | Uye | Uye | Mukellef seklinde 4 imza blogu."""
    aralik = [(1, 1), (2, 3), (4, 5), (6, 6)]
    unvanlar = ["Başkan", "Üye", "Üye", "Mükellef"]
    isimler = []
    for u in unvanlar[:3]:
        hedef = _unvan_normal(u)
        uygun = next((k["ad_soyad"] for k in imzalayanlar
                       if _unvan_normal(k.get("gorev") or k.get("unvan")) == hedef
                       and k["ad_soyad"] not in isimler), None)
        if uygun is None:
            uygun = next((k["ad_soyad"] for k in imzalayanlar if k["ad_soyad"] not in isimler), "")
        isimler.append(uygun)
    isimler.append(mukellef_adi)

    for (c1, c2), isim in zip(aralik, isimler):
        if c1 != c2:
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1, value=isim)
        cell.alignment = SIGDIR
        cell.font = FONT_NORMAL

    for (c1, c2), unvan in zip(aralik, unvanlar):
        if c1 != c2:
            ws.merge_cells(start_row=row + 1, start_column=c1, end_row=row + 1, end_column=c2)
        cell = ws.cell(row=row + 1, column=c1, value=unvan)
        cell.alignment = CENTER
        cell.font = FONT_BOLD

    return row + 2


def _imza_bloklari_3lu(ws, row, imzalayanlar):
    """Baskan | Uye | Uye seklinde 3 imza blogu (Gelmedi tutanagi)."""
    aralik = [(1, 2), (3, 4), (5, 6)]
    unvanlar = ["Başkan", "Üye", "Üye"]
    isimler = []
    for u in unvanlar:
        hedef = _unvan_normal(u)
        uygun = next((k["ad_soyad"] for k in imzalayanlar
                       if _unvan_normal(k.get("gorev") or k.get("unvan")) == hedef
                       and k["ad_soyad"] not in isimler), None)
        if uygun is None:
            uygun = next((k["ad_soyad"] for k in imzalayanlar if k["ad_soyad"] not in isimler), "")
        isimler.append(uygun)

    for (c1, c2), isim in zip(aralik, isimler):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row=row, column=c1, value=isim)
        cell.alignment = SIGDIR
        cell.font = FONT_NORMAL

    for (c1, c2), unvan in zip(aralik, unvanlar):
        ws.merge_cells(start_row=row + 1, start_column=c1, end_row=row + 1, end_column=c2)
        cell = ws.cell(row=row + 1, column=c1, value=unvan)
        cell.alignment = CENTER
        cell.font = FONT_BOLD

    return row + 2


def _teslim_alma_ibaresi(ws, row, mukellef_adi):
    """Madde 10: mukellefin tutanagin bir nushasini aldigina dair ibare.

    Mukellef imza sutunu hizasinda (D:F) yazilir; ibare, imza boslugu ve
    teslim alan mukellefin adi alt alta gelir.
    """
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    c = ws.cell(row=row, column=4,
                value="Bu tutanağın bir nüshasını uzlaşma komisyonunda\n"
                      "....../....../.......... tarihinde aldım.")
    c.font = FONT_NORMAL
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 44
    ws.row_dimensions[row + 1].height = 28  # imza boslugu
    ws.merge_cells(start_row=row + 2, start_column=4, end_row=row + 2, end_column=6)
    c2 = ws.cell(row=row + 2, column=4, value=mukellef_adi)
    c2.font = FONT_BOLD
    c2.alignment = CENTER
    return row + 3


# ---------------------------------------------------------------------------
# Sablon 1 ve 2: UZLASMA TUTANAGI (uzlasildi / uzlasilamadi)
# ---------------------------------------------------------------------------

def uzlasma_tutanagi_olustur(dosya_yolu, kurum, tutanak_no, toplanti_tarih_saat, mukellef,
                              kalemler, imzalayanlar, sonuc, davet_tarih_saat=""):
    """sonuc: 'uzlasildi' veya 'uzlasilamadi'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tutanak"
    _kur_kolon_genislikleri(ws)

    baslik = "UZLAŞMA TUTANAĞI" if sonuc == "uzlasildi" else "UZLAŞMA VAKİ OLMADI TUTANAĞI"
    row = _baslik_blogu(ws, kurum, baslik)
    tarih_metni = toplanti_tarih_saat
    if (davet_tarih_saat or "").strip():
        tarih_metni = f"{toplanti_tarih_saat} / {davet_tarih_saat.strip()}"
    row = _ust_bilgi_blogu(ws, row, kurum, tutanak_no, tarih_metni, mukellef)

    tarih_str, saat_str = _tarih_saat_ayir(toplanti_tarih_saat)
    if sonuc == "uzlasildi":
        aciklama = (
            f"Aşağıda isim ve ünvanları yazılı Başkan ve üyelerinden teşekkül eden Uzlaşma "
            f"Komisyonumuz mükellefin iştirakiyle {tarih_str} tarihinde saat {saat_str}'da "
            f"toplanarak tabloda yazılı vergi ve cezalar ile önerilen tutarlar üzerinde "
            f"uzlaşma sağlanmıştır."
        )
        tutar_basligi = "UZLAŞILAN TUTAR"
    else:
        aciklama = (
            f"Aşağıda isim ve ünvanları yazılı Başkan ve üyelerinden teşekkül eden Uzlaşma "
            f"Komisyonumuz mükellefin iştirakiyle {tarih_str} tarihinde saat {saat_str}'da "
            f"toplanarak tabloda yazılı vergi ve cezalar ile önerilen tutarlar üzerinde "
            f"uzlaşma vaki olmamıştır.\n\n"
            f"     Uzlaşma yönetmeliğinin 10. maddesine göre mükellefin önerilen bu miktarları "
            f"dava açma süresinin son günü akşamına kadar kabul ettiğini bildiren bir dilekçe "
            f"ile başvurması halinde uzlaşma vaki olmuş sayılacaktır."
        )
        tutar_basligi = "ÖNERİLEN TUTAR"

    row = _paragraf(ws, row, aciklama)

    # Tablo basligi (2 satir, birlesik)
    baslik_satiri = row
    basliklar = ["İhbarname Numarası", "Vergi Türü", "Cezanın Türü", "Vergi ve Cezanın Miktarı"]
    for col, metin in enumerate(basliklar, start=1):
        ws.merge_cells(start_row=baslik_satiri, start_column=col, end_row=baslik_satiri + 1, end_column=col)
        _hucre(ws, baslik_satiri, col, metin, font=FONT_BOLD)
    ws.merge_cells(start_row=baslik_satiri, start_column=5, end_row=baslik_satiri, end_column=6)
    _hucre(ws, baslik_satiri, 5, tutar_basligi, font=FONT_BOLD)
    _hucre(ws, baslik_satiri + 1, 5, "Rakamla (TL)", font=FONT_BOLD)
    _hucre(ws, baslik_satiri + 1, 6, "Yazıyla (TL)", font=FONT_BOLD)
    row = baslik_satiri + 2

    toplam_miktar = 0.0
    toplam_uzlasilan = 0.0
    for kalem in kalemler:
        miktar = float(kalem["miktar"])
        uzlasilan = float(kalem["uzlasilan_tutar"])
        toplam_miktar += miktar
        toplam_uzlasilan += uzlasilan
        _hucre(ws, row, 1, kalem["fis_no"], align=SIGDIR)
        _hucre(ws, row, 2, (kalem.get("vergi_turu_kod") or "").strip(), align=SIGDIR)
        _hucre(ws, row, 3, (kalem.get("ceza_kodu") or "").strip(), align=SIGDIR)
        _hucre(ws, row, 4, miktar, num_format="#,##0.00", align=SIGDIR)
        _hucre(ws, row, 5, uzlasilan, num_format="#,##0.00", align=SIGDIR)
        yazi = tutar_yaziya(uzlasilan)
        _hucre(ws, row, 6, yazi, align=JUSTIFY)
        ws.row_dimensions[row].height = max(21, _metin_satir_sayisi(yazi, 22) * 15 + 6)
        row += 1

    toplam_miktar = round(toplam_miktar, 2)
    toplam_uzlasilan = round(toplam_uzlasilan, 2)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _hucre(ws, row, 1, "TOPLAM", font=FONT_BOLD)
    _hucre(ws, row, 4, toplam_miktar, font=FONT_BOLD, num_format="#,##0.00", align=SIGDIR)
    _hucre(ws, row, 5, toplam_uzlasilan, font=FONT_BOLD, num_format="#,##0.00", align=SIGDIR)
    toplam_yazi = tutar_yaziya(toplam_uzlasilan)
    _hucre(ws, row, 6, toplam_yazi, font=FONT_BOLD, align=JUSTIFY)
    ws.row_dimensions[row].height = max(21, _metin_satir_sayisi(toplam_yazi, 22) * 15 + 6)
    _tablo_kenarligi_tamamla(ws, baslik_satiri, row)
    _baslik_satiri_boya(ws, baslik_satiri, baslik_satiri + 1)
    row += 1

    row = _paragraf(
        ws, row,
        "     İşbu Uzlaşma Komisyonu Tutanağı (3) nüsha tanzim edilerek okundu. Doğruluğu "
        "anlaşılarak mükellefle birlikte müştereken imzalandı. Düzenlenen tutanağın bir "
        "örneği mükellefe komisyonda verildi.",
    )
    if sonuc == "uzlasildi":
        row = _paragraf(
            ws, row,
            "     NOT: İşbu uzlaşılan vergiler için V.U.K.nun 112.maddesi 3.fıkrası gereğince "
            "normal vade tarihinden itibaren uzlaşma tutanağının imzalandığı tarihe kadar geçen "
            "zaman için ayrıca Vergi Dairesince gecikme faizi hesaplanacaktır.",
        )
    else:
        # Uzlasmanin vaki olmamasi: dava acma suresi notu (Madde 13)
        row = _paragraf(ws, row, "     " + DAVA_SURESI_NOTU)
    # Islak imza icin isimlerin ustunde bos alan birakilir
    ws.row_dimensions[row].height = 30
    ws.row_dimensions[row + 1].height = 30
    row += 2

    row = _imza_bloklari_4lu(ws, row, imzalayanlar, mukellef.get("ad_unvan", ""))
    # Madde 10: teslim alma ibaresi, mukellef isminin iki satir altina
    _teslim_alma_ibaresi(ws, row, mukellef.get("ad_unvan", ""))

    _sayfa_ayari(ws)
    os.makedirs(os.path.dirname(dosya_yolu), exist_ok=True)
    wb.save(dosya_yolu)
    return dosya_yolu


# ---------------------------------------------------------------------------
# Sablon 3: UZLASMA KOMISYON KARAR TUTANAGI (gelmedi)
# ---------------------------------------------------------------------------

def gelmeme_tutanagi_olustur(dosya_yolu, kurum, tutanak_no, toplanti_tarih_saat,
                              davet_tarih_saat, mukellef, kalemler, imzalayanlar):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tutanak"
    _kur_kolon_genislikleri(ws)

    row = _baslik_blogu(ws, kurum, "UZLAŞMA TEMİN EDİLEMEDİ TUTANAĞI")

    tarih_metni = f"{toplanti_tarih_saat} / {davet_tarih_saat}"
    row = _ust_bilgi_blogu(ws, row, kurum, tutanak_no, tarih_metni, mukellef,
                            ara_baslik="UZLAŞMA TALEP EDENİN")

    row += 1  # buyuk harfli baslik ile ustteki satir arasinda bosluk
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = ws.cell(row=row, column=1, value="UZLAŞMA TALEP EDİLEN")
    c.font = FONT_BOLD
    c.alignment = CENTER_TITLE
    row += 1

    baslik_satiri = row
    basliklar = ["İhbarname Numarası", "Vergi Türü", "Cezanın Türü", "Vergi ve Cezanın Miktarı"]
    for col, metin in enumerate(basliklar, start=1):
        if col == 4:
            ws.merge_cells(start_row=baslik_satiri, start_column=4, end_row=baslik_satiri, end_column=6)
        _hucre(ws, baslik_satiri, col, metin, font=FONT_BOLD)
    row = baslik_satiri + 1

    toplam_miktar = 0.0
    for kalem in kalemler:
        miktar = float(kalem["miktar"])
        toplam_miktar += miktar
        _hucre(ws, row, 1, kalem["fis_no"], align=SIGDIR)
        _hucre(ws, row, 2, (kalem.get("vergi_turu_kod") or "").strip(), align=SIGDIR)
        _hucre(ws, row, 3, (kalem.get("ceza_kodu") or "").strip(), align=SIGDIR)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        _hucre(ws, row, 4, miktar, num_format="#,##0.00", align=SIGDIR)
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _hucre(ws, row, 1, "TOPLAM", font=FONT_BOLD)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    _hucre(ws, row, 4, toplam_miktar, font=FONT_BOLD, num_format="#,##0.00", align=SIGDIR)
    _tablo_kenarligi_tamamla(ws, baslik_satiri, row)
    _baslik_satiri_boya(ws, baslik_satiri, baslik_satiri)
    row += 1

    toplanti_tarih, toplanti_saat = _tarih_saat_ayir(toplanti_tarih_saat)
    davet_tarih, davet_saat = _tarih_saat_ayir(davet_tarih_saat)
    vergi_dairesi = kurum.get("vergi_dairesi", "")
    aciklama = (
        f"Uzlaşma Komisyonumuz 22/10/2005 tarih ve 25974 sayılı Uzlaşma Yönetmeliğinde "
        f"Değişiklik Yapılmasına Dair Yönetmeliğin 6. maddesine uygun olarak aşağıda isim ve "
        f"unvanları yazılı Başkan ve üyelerinden teşekkül eden Uzlaşma Komisyonumuz "
        f"{toplanti_tarih} tarihinde saat {toplanti_saat}'da toplandı.\n\n"
        f"Ancak uzlaşma isteminde bulunan ve yukarıda açık kimliği belirtilen mükellefe "
        f"uzlaşma görüşmesinin {davet_tarih} tarihinde saat {davet_saat}'da {vergi_dairesi}'nde "
        f"yapılacağına ilişkin uzlaşma davetiyesi usulüne uygun tebliğ edildiği halde, uzlaşma "
        f"davetiyesinde tayin edilen gün ve saatte bizzat veya vekili vasıtasıyla Uzlaşma "
        f"Komisyonuna gelmediğinden uzlaşma temin edilememiştir. Durumu tespit eden bu tutanak "
        f"komisyon üyelerince okunduktan sonra müştereken imza altına alındı."
    )
    row = _paragraf(ws, row, aciklama)
    # Uzlasmanin temin edilememesi: dava acma suresi notu (Madde 13)
    row = _paragraf(ws, row, "     " + DAVA_SURESI_NOTU)
    # Islak imza icin isimlerin ustunde bos alan birakilir
    ws.row_dimensions[row].height = 30
    ws.row_dimensions[row + 1].height = 30
    row += 2

    _imza_bloklari_3lu(ws, row, imzalayanlar)

    _sayfa_ayari(ws)
    os.makedirs(os.path.dirname(dosya_yolu), exist_ok=True)
    wb.save(dosya_yolu)
    return dosya_yolu


def _tarih_saat_ayir(tarih_saat_str):
    """'DD.MM.YYYY HH:MM' -> ('DD.MM.YYYY', 'HH:MM')."""
    if not tarih_saat_str:
        return "", ""
    parcalar = tarih_saat_str.strip().split(" ")
    if len(parcalar) >= 2:
        return parcalar[0], parcalar[1]
    return parcalar[0], ""


def tutanak_olustur_excel(dosya_yolu, kurum, tutanak_no, toplanti_tarih_saat, davet_tarih_saat,
                           mukellef, kalemler, imzalayanlar, sonuc):
    """Sonuca gore uygun sablonu uretir."""
    if sonuc == "gelmedi":
        return gelmeme_tutanagi_olustur(
            dosya_yolu, kurum, tutanak_no, toplanti_tarih_saat, davet_tarih_saat,
            mukellef, kalemler, imzalayanlar,
        )
    return uzlasma_tutanagi_olustur(
        dosya_yolu, kurum, tutanak_no, toplanti_tarih_saat, mukellef, kalemler, imzalayanlar,
        sonuc, davet_tarih_saat,
    )


# ---------------------------------------------------------------------------
# Aylik huzur hakki puantaj cetveli
# ---------------------------------------------------------------------------

import calendar

AY_ADLARI = ["", "OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN",
             "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]

MEDIUM = Side(style="medium")
BORDER_MEDIUM = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

P_FONT_BASLIK = Font(name="Times New Roman", size=16, bold=True)
P_FONT_BOLD12 = Font(name="Times New Roman", size=12, bold=True)
P_FONT_NORMAL = Font(name="Times New Roman", size=11)
P_FONT_BOLD11 = Font(name="Times New Roman", size=11, bold=True)

SOL_ORTA = Alignment(horizontal="left", vertical="center")


def _p_hucre(ws, row, col, deger, font, hiza=CENTER, kenarlik=BORDER_MEDIUM):
    c = ws.cell(row=row, column=col, value=deger)
    c.font = font
    c.alignment = hiza
    if kenarlik is not None:
        c.border = kenarlik
    return c


def _p_blok_kenarlik(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER_MEDIUM


def puantaj_cetveli_olustur(dosya_yolu, kurum, yil, ay, uyeler):
    """Aylik huzur hakki puantaj cetveli.

    uyeler: [{"ad_soyad", "unvan", "gunler": iterable[int], "oturum_sayisi": int}]
    Gun sutunu sayisi ayin gercek gun sayisina gore belirlenir (28-31).
    """
    gun_sayisi = calendar.monthrange(yil, ay)[1]
    son_gun_kolonu = 3 + gun_sayisi          # D'den itibaren gun kolonlari
    toplam_kolonu = son_gun_kolonu + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Puantaj Cetveli"

    ws.column_dimensions["A"].width = 4.29
    ws.column_dimensions["B"].width = 30.7
    ws.column_dimensions["C"].width = 15.71
    for c in range(4, son_gun_kolonu + 1):
        ws.column_dimensions[get_column_letter(c)].width = 4.29
    ws.column_dimensions[get_column_letter(toplam_kolonu)].width = 18.71

    # 1. satir: baslik
    ws.row_dimensions[1].height = 60
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=toplam_kolonu)
    _p_hucre(ws, 1, 1, "TARHİYAT SONRASI UZLAŞMA ÜYELERİ HUZUR HAKKI PUANTAJ CETVELİ",
             P_FONT_BASLIK, kenarlik=None)

    # 2. satir: daire ve donem (A kolonundan baslar ki sol ust kose cerceveli olsun)
    ws.row_dimensions[2].height = 30
    donem_baslik_kolonu = son_gun_kolonu - 7           # ornekte Z2:AD2 (5 kolon)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=donem_baslik_kolonu - 1)
    ws.merge_cells(start_row=2, start_column=donem_baslik_kolonu, end_row=2,
                   end_column=donem_baslik_kolonu + 4)
    ws.merge_cells(start_row=2, start_column=donem_baslik_kolonu + 5, end_row=2,
                   end_column=toplam_kolonu)
    _p_hucre(ws, 2, 1, f"Dairesi : {kurum.get('vergi_dairesi', '')}", P_FONT_BOLD12, SOL_ORTA)
    _p_hucre(ws, 2, donem_baslik_kolonu, "Dönemi", P_FONT_BOLD12)
    _p_hucre(ws, 2, donem_baslik_kolonu + 5, f"{AY_ADLARI[ay]} / {yil}", P_FONT_BOLD12)
    _p_blok_kenarlik(ws, 2, 1, 2, toplam_kolonu)

    # 3-4. satirlar: tablo basligi
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 30
    ws.merge_cells("A3:A4")
    ws.merge_cells("B3:C3")
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=son_gun_kolonu)
    ws.merge_cells(start_row=3, start_column=toplam_kolonu, end_row=4, end_column=toplam_kolonu)
    _p_hucre(ws, 3, 2, "KOMİSYON ÜYELERİ", P_FONT_BOLD12)
    _p_hucre(ws, 3, 4, "OTURUM GÜNLERİ", P_FONT_BOLD12)
    _p_hucre(ws, 3, toplam_kolonu, "TOPLAM OTURUM SAYISI", P_FONT_BOLD12)
    _p_hucre(ws, 4, 2, "ADI - SOYADI", P_FONT_BOLD12)
    _p_hucre(ws, 4, 3, "UNVANI", P_FONT_BOLD12)
    for gun in range(1, gun_sayisi + 1):
        _p_hucre(ws, 4, 3 + gun, gun, P_FONT_BOLD11)
    _p_blok_kenarlik(ws, 3, 1, 4, toplam_kolonu)

    # 5+ satirlar: uyeler
    row = 5
    for sira, uye in enumerate(uyeler, start=1):
        ws.row_dimensions[row].height = 30
        _p_hucre(ws, row, 1, sira, P_FONT_NORMAL)
        _p_hucre(ws, row, 2, uye.get("ad_soyad", ""), P_FONT_NORMAL)
        _p_hucre(ws, row, 3, uye.get("unvan", ""), P_FONT_NORMAL)
        gunler = set(uye.get("gunler") or [])
        for gun in range(1, gun_sayisi + 1):
            if gun in gunler:
                _p_hucre(ws, row, 3 + gun, "X", P_FONT_BOLD11)
        _p_hucre(ws, row, toplam_kolonu, uye.get("oturum_sayisi", len(gunler)), P_FONT_NORMAL)
        _p_blok_kenarlik(ws, row, 1, row, toplam_kolonu)
        row += 1

    # alt blok: onay metni ve imzalar (metinler dar gun kolonlarinda kesilmesin
    # diye her biri birkac kolonu kapsayan birlesik hucrelere yazilir)
    def _genis_hucre(r, c1, c2, deger):
        c2 = min(c2, toplam_kolonu)
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        _p_hucre(ws, r, c1, deger, P_FONT_BOLD12, CENTER_TITLE, kenarlik=None)

    onay_kolonu = toplam_kolonu - 6                    # ornekte AB (34-6=28)
    row += 1
    _genis_hucre(row, onay_kolonu - 2, onay_kolonu + 6, "Kayıtlarımıza Uygundur")
    row += 1
    from datetime import date
    bugun = date.today()
    _genis_hucre(row, onay_kolonu - 2, onay_kolonu + 6, f"{bugun.day}.{bugun.month}.{bugun.year}")

    row += 4
    imza_kolonlari = [(3, "Memur", "Düzenleyen"), (14, "Şef", "Kontrol Eden"),
                      (onay_kolonu, "Vergi Dairesi Müdürü", "Onay")]
    for kolon, unvan, gorev in imza_kolonlari:
        c1, c2 = (kolon, kolon + 1) if kolon == 3 else (kolon - 2, kolon + 6)
        _genis_hucre(row, c1, c2, "...........................")
        _genis_hucre(row + 1, c1, c2, unvan)
        _genis_hucre(row + 2, c1, c2, gorev)

    # yazdirma ayarlari: A4 yatay, tek sayfaya sigdir
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.print_area = f"A1:{get_column_letter(toplam_kolonu)}{row + 2}"

    wb.save(dosya_yolu)
    return dosya_yolu
